import os
import copy
import unicodedata
import pandas as pd
import tkinter as tk
from tkinter import Tk, filedialog, messagebox

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================================================
# CONFIGURACIÓN
# =========================================================

FILA_ENCABEZADO = 10
FILA_PRIMER_ALUMNO = 11

COL_ALUMNO = 1
COL_TP1 = 2
COL_TP2 = 3
COL_PRM1 = 4
COL_INAS = 5

AMARILLO = PatternFill(
    start_color="FFFF00",
    end_color="FFFF00",
    fill_type="solid"
)

GRIS_ENCABEZADO = PatternFill(
    start_color="D9D9D9",
    end_color="D9D9D9",
    fill_type="solid"
)

BORDE_FINO = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


# =========================================================
# FUNCIONES AUXILIARES
# =========================================================

def convertir_nota(valor):
    if valor is None:
        return 0

    texto = str(valor).strip().upper()

    if texto == "" or texto == "A":
        return 0

    try:
        return float(texto.replace(",", "."))
    except:
        return 0


def limpiar_dni(texto):
    if texto is None:
        return ""

    texto = str(texto)
    texto = texto.replace("DNT", "DNI")
    texto = texto.replace("DNI", "\nDNI ")
    texto = texto.replace("\nDNI  ", "\nDNI ")

    return texto.strip()


def normalizar_texto(texto):
    """
    Convierte texto a mayúsculas y elimina tildes.
    Ejemplo: GARCÍA -> GARCIA
    """
    texto = str(texto).upper().strip()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(
        caracter for caracter in texto
        if unicodedata.category(caracter) != "Mn"
    )
    return texto


def obtener_primer_apellido(alumno):
    """
    Detecta el primer apellido normalizado.
    Ejemplo:
    'GARCÍA LOPEZ, JUAN\nDNI 12345678' -> 'GARCIA'
    'GARCIA PEREZ, ANA\nDNI 87654321' -> 'GARCIA'
    """
    if alumno is None:
        return ""

    texto = normalizar_texto(alumno)

    if not texto:
        return ""

    parte_apellidos = texto.split(",")[0].strip()

    if not parte_apellidos:
        return ""

    primer_apellido = parte_apellidos.split()[0]

    return primer_apellido


def ajustar_ancho_columna(ws, col, minimo=10, maximo=60):
    letra = get_column_letter(col)
    max_len = 0

    for cell in ws[letra]:
        if cell.value is not None:
            largo = max(
                len(linea)
                for linea in str(cell.value).split("\n")
            )
            max_len = max(max_len, largo)

    ws.column_dimensions[letra].width = min(
        max(max_len + 2, minimo),
        maximo
    )


def crear_hoja_lib(wb, alumnos_libres):
    if "LIB" in wb.sheetnames:
        del wb["LIB"]

    ws = wb.create_sheet("LIB")

    encabezados = ["ALUMNO", "TP1", "TP2", "PRM1"]

    for col, titulo in enumerate(encabezados, start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = GRIS_ENCABEZADO
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.border = BORDE_FINO

    fila = 2

    for alumno in alumnos_libres:
        ws.cell(fila, 1, alumno["alumno"])
        ws.cell(fila, 2, alumno["tp1"])
        ws.cell(fila, 3, alumno["tp2"])
        ws.cell(fila, 4, alumno["prm1"])

        for col in range(1, 5):
            c = ws.cell(fila, col)
            c.fill = AMARILLO
            c.border = BORDE_FINO
            c.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

        fila += 1

    ajustar_ancho_columna(ws, 1, minimo=35, maximo=70)
    ajustar_ancho_columna(ws, 2)
    ajustar_ancho_columna(ws, 3)
    ajustar_ancho_columna(ws, 4)

    ws.freeze_panes = "A2"


def crear_hoja_reg_2et(wb, regulares):
    if "REG 2ET" in wb.sheetnames:
        del wb["REG 2ET"]

    ws = wb.create_sheet("REG 2ET")
    ws["A1"] = regulares
    ws["A1"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 12


def crear_hoja_inas(wb):
    if "INAS" in wb.sheetnames:
        del wb["INAS"]

    wb.create_sheet("INAS")


# =========================================================
# PROCESAR XLSX
# =========================================================

def procesar_xlsx(ruta_archivo):
    wb = load_workbook(ruta_archivo)
    ws = wb.active
    ws.title = "REPORTE"

    conteo_apellidos = {}

    # Limpiar columnas D y E originales
    for fila in range(1, ws.max_row + 1):
        ws.cell(fila, COL_PRM1).value = None
        ws.cell(fila, COL_INAS).value = None

    # Encabezados finales
    ws.cell(FILA_ENCABEZADO, COL_ALUMNO).value = "ALUMNO"
    ws.cell(FILA_ENCABEZADO, COL_TP1).value = "TP1"
    ws.cell(FILA_ENCABEZADO, COL_TP2).value = "TP2"
    ws.cell(FILA_ENCABEZADO, COL_PRM1).value = "PRM1"
    ws.cell(FILA_ENCABEZADO, COL_INAS).value = "INAS"

    for col in range(1, 6):
        cell = ws.cell(FILA_ENCABEZADO, col)
        cell.fill = GRIS_ENCABEZADO
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.border = BORDE_FINO

    alumnos_libres = []
    regulares = 0
    libres = 0

    fila = FILA_PRIMER_ALUMNO

    while True:
        alumno = ws.cell(fila, COL_ALUMNO).value

        if alumno is None or str(alumno).strip() == "":
            break

        alumno_limpio = limpiar_dni(alumno)
        ws.cell(fila, COL_ALUMNO).value = alumno_limpio

        apellido = obtener_primer_apellido(alumno_limpio)
        if apellido:
            conteo_apellidos[apellido] = conteo_apellidos.get(apellido, 0) + 1

        tp1 = convertir_nota(ws.cell(fila, COL_TP1).value)
        tp2 = convertir_nota(ws.cell(fila, COL_TP2).value)
        prm1 = round((tp1 + tp2) / 2, 2)

        ws.cell(fila, COL_TP1).value = tp1
        ws.cell(fila, COL_TP2).value = tp2
        ws.cell(fila, COL_PRM1).value = prm1
        ws.cell(fila, COL_INAS).value = ""

        for col in range(1, 6):
            ws.cell(fila, col).border = BORDE_FINO
            ws.cell(fila, col).alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

        if prm1 < 4:
            libres += 1

            alumnos_libres.append({
                "alumno": alumno_limpio,
                "tp1": tp1,
                "tp2": tp2,
                "prm1": prm1
            })

            for col in range(1, 6):
                ws.cell(fila, col).fill = AMARILLO
        else:
            regulares += 1

        fila += 1

    ultima_fila = fila - 1
    ws.auto_filter.ref = f"A{FILA_ENCABEZADO}:E{ultima_fila}"

    ajustar_ancho_columna(ws, 1, minimo=35, maximo=70)

    for col in range(2, 6):
        ajustar_ancho_columna(ws, col, minimo=10, maximo=18)

    crear_hoja_lib(wb, alumnos_libres)
    crear_hoja_reg_2et(wb, regulares)
    crear_hoja_inas(wb)

    apellidos_repetidos = [
        apellido
        for apellido, cantidad in conteo_apellidos.items()
        if cantidad > 1
    ]

    return wb, regulares, libres, apellidos_repetidos


# =========================================================
# PROCESAR ODS
# =========================================================

def procesar_ods(ruta_archivo):
    df = pd.read_excel(
        ruta_archivo,
        header=None,
        engine="odf"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "REPORTE"

    conteo_apellidos = {}

    # Copiar filas 1 a 9
    for r in range(1, FILA_ENCABEZADO):
        for c in range(1, 6):
            valor = None

            try:
                valor = df.iloc[r - 1, c - 1]

                if pd.isna(valor):
                    valor = None

            except:
                valor = None

            ws.cell(r, c, valor)

    # Encabezados tabla
    encabezados = ["ALUMNO", "TP1", "TP2", "PRM1", "INAS"]

    for col, titulo in enumerate(encabezados, start=1):
        cell = ws.cell(FILA_ENCABEZADO, col, titulo)
        cell.fill = GRIS_ENCABEZADO
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.border = BORDE_FINO

    alumnos_libres = []
    regulares = 0
    libres = 0

    fila_excel = FILA_PRIMER_ALUMNO
    fila_df = FILA_PRIMER_ALUMNO - 1

    while fila_df < len(df):
        alumno = df.iloc[fila_df, 0]

        if pd.isna(alumno) or str(alumno).strip() == "":
            break

        alumno_limpio = limpiar_dni(alumno)

        apellido = obtener_primer_apellido(alumno_limpio)
        if apellido:
            conteo_apellidos[apellido] = conteo_apellidos.get(apellido, 0) + 1

        tp1 = convertir_nota(df.iloc[fila_df, 1])
        tp2 = convertir_nota(df.iloc[fila_df, 2])
        prm1 = round((tp1 + tp2) / 2, 2)

        valores = [
            alumno_limpio,
            tp1,
            tp2,
            prm1,
            ""
        ]

        for col, valor in enumerate(valores, start=1):
            cell = ws.cell(fila_excel, col, valor)
            cell.border = BORDE_FINO
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

        if prm1 < 4:
            libres += 1

            alumnos_libres.append({
                "alumno": alumno_limpio,
                "tp1": tp1,
                "tp2": tp2,
                "prm1": prm1
            })

            for col in range(1, 6):
                ws.cell(fila_excel, col).fill = AMARILLO
        else:
            regulares += 1

        fila_excel += 1
        fila_df += 1

    ultima_fila = fila_excel - 1
    ws.auto_filter.ref = f"A{FILA_ENCABEZADO}:E{ultima_fila}"

    ws.column_dimensions["A"].width = 45

    for col in range(2, 6):
        ws.column_dimensions[get_column_letter(col)].width = 12

    crear_hoja_lib(wb, alumnos_libres)
    crear_hoja_reg_2et(wb, regulares)
    crear_hoja_inas(wb)

    apellidos_repetidos = [
        apellido
        for apellido, cantidad in conteo_apellidos.items()
        if cantidad > 1
    ]

    return wb, regulares, libres, apellidos_repetidos


# =========================================================
# PROGRAMA PRINCIPAL
# =========================================================

def main():
    Tk().withdraw()

    ruta_archivo = filedialog.askopenfilename(
        title="Seleccionar archivo de Guaraní",
        filetypes=[
            ("Archivos compatibles", "*.xlsx *.ods"),
            ("Excel", "*.xlsx"),
            ("OpenDocument", "*.ods")
        ]
    )

    if not ruta_archivo:
        return

    extension = os.path.splitext(ruta_archivo)[1].lower()

    try:
        if extension == ".xlsx":
            wb, regulares, libres, apellidos_repetidos = procesar_xlsx(ruta_archivo)

        elif extension == ".ods":
            wb, regulares, libres, apellidos_repetidos = procesar_ods(ruta_archivo)

        else:
            messagebox.showerror(
                "Formato no compatible",
                "Usá un archivo .xlsx u .ods.\n\n"
                "Si el archivo está en .xls, abrilo en Excel y guardalo como .xlsx."
            )
            return

        nombre_base = os.path.splitext(
            os.path.basename(ruta_archivo)
        )[0]

        salida = os.path.join(
            os.path.dirname(ruta_archivo),
            f"PROCESADO_{nombre_base}.xlsx"
        )

        wb.save(salida)

        messagebox.showinfo(
            "Proceso finalizado",
            f"Archivo creado correctamente.\n\n"
            f"Regulares para REG 2ET: {regulares}\n"
            f"Libres: {libres}\n\n"
            f"Archivo:\n{salida}"
        )

        try:
            os.startfile(os.path.dirname(salida))
        except:
            pass

        if apellidos_repetidos:
            lista = "\n".join(apellidos_repetidos)

            messagebox.showwarning(
                "ATENCIÓN",
                "ATENCIÓN! MATERIA CON ALUMNOS CON MISMO APELLIDO.\n\n"
                "Revisar especialmente el orden al cargar o copiar inasistencias.\n\n"
                f"Apellidos repetidos detectados:\n{lista}"
            )

    except Exception as e:
        messagebox.showerror(
            "Error",
            f"Ocurrió un error:\n\n{e}"
        )


if __name__ == "__main__":
    main()
