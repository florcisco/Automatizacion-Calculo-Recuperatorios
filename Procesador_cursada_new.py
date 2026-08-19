# ============================================================
# PROCESADOR DE CURSADAS
# Versión 4.0
#
# Autor: Francisco Lombroni



# ============================================================
# MÓDULO 0
# VENTANA DE PROGRESO
# ============================================================

import tkinter as tk
from tkinter import messagebox, simpledialog, ttk

from pandas.io import excel

ventana_progreso = None
barra_progreso = None
etiqueta_estado = None


def crear_ventana_progreso():
    """
    Crea la ventana de progreso del programa.
    """

    global ventana_progreso
    global barra_progreso
    global etiqueta_estado

    ventana_progreso = tk.Tk()

    ventana_progreso.title(TITULO)

    ventana_progreso.geometry("500x180")

    ventana_progreso.resizable(False, False)

    ventana_progreso.attributes("-topmost", True)

    tk.Label(
        ventana_progreso,
        text="Procesador de Cursadas",
        font=("Segoe UI", 12, "bold")
    ).pack(pady=(15, 10))

    etiqueta_estado = tk.Label(
        ventana_progreso,
        text="Inicializando...",
        font=("Segoe UI", 10)
    )

    etiqueta_estado.pack()

    barra_progreso = ttk.Progressbar(
        ventana_progreso,
        orient="horizontal",
        mode="determinate",
        length=420
    )

    barra_progreso.pack(pady=20)

    barra_progreso["maximum"] = 100

    barra_progreso["value"] = 0

    ttk.Label(
        ventana_progreso,
        text="Espere por favor...",
        font=("Segoe UI", 9)
    ).pack()

    ventana_progreso.update()

def mostrar_error(mensaje):

    cerrar_ventana_progreso()

    messagebox.showerror(
        TITULO,
        mensaje
    )


def mostrar_advertencia(mensaje):

    messagebox.showwarning(
        TITULO,
        mensaje
    )


def mostrar_info(mensaje):

    messagebox.showinfo(
        TITULO,
        mensaje
    )
  
def actualizar_progreso(mensaje, porcentaje):
    """
    Actualiza el mensaje y la barra de progreso.
    """

    etiqueta_estado.config(text=mensaje)

    barra_progreso["value"] = porcentaje

    ventana_progreso.update_idletasks()


def cerrar_ventana_progreso():
    """
    Cierra la ventana de progreso si todavía existe.
    """

    global ventana_progreso

    if ventana_progreso is None:
        return

    try:
        if ventana_progreso.winfo_exists():
            ventana_progreso.destroy()
    except tk.TclError:
        pass

    ventana_progreso = None


# ============================================================
# MÓDULO 0.5
# FUNCIONES AUXILIARES
# ============================================================

def obtener_lista_alumnos(df, columna_alumno="ALUMNO"):
    """
    Obtiene los alumnos de un DataFrame.

    La clave utilizada para identificar a cada alumno
    es el nombre completo normalizado.

    También registra los apellidos repetidos,
    ignorando mayúsculas, minúsculas y tildes.
    """

    global apellidos_repetidos

    alumnos = {}
    apellidos_encontrados = set()

    for indice, valor in df[columna_alumno].items():

        if pd.isna(valor):
            continue

        nombre = str(valor).strip()

        if nombre == "":
            continue

        # ------------------------------------------------
        # CLAVE DEL ALUMNO
        # ------------------------------------------------

        clave = obtener_clave_alumno(
            nombre
        )

        # ------------------------------------------------
        # DETECTAR APELLIDO REPETIDO
        # ------------------------------------------------

        apellido = normalizar_texto(
            obtener_apellido(nombre)
        )

        if apellido in apellidos_encontrados:

            if apellido not in apellidos_repetidos:

                apellidos_repetidos.append(
                    apellido
                )

        else:

            apellidos_encontrados.add(
                apellido
            )

        # ------------------------------------------------
        # GUARDAR ALUMNO
        # ------------------------------------------------

        alumnos[clave] = {
            "fila": indice,
            "nombre": nombre
        }

    return alumnos

def mostrar_alerta_apellidos_repetidos():
    """
    Muestra una alerta si se detectaron apellidos repetidos.

    El programa continúa normalmente.
    """

    if not apellidos_repetidos:
        return

    apellidos = sorted(
        set(apellidos_repetidos)
    )

    mensaje = (
        "ATENCIÓN: MATERIA CON ALUMNOS "
        "CON MISMO APELLIDO.\n\n"
        "Se detectaron los siguientes "
        "apellidos repetidos:\n\n"
        + "\n".join(
            f"• {apellido}"
            for apellido in apellidos
        )
    )

    messagebox.showwarning(
        TITULO,
        mensaje
    )
    
def quitar_tildes(texto):
    """
    Elimina las tildes de un texto.
    """

    if texto is None:
        return ""

    texto = str(texto)

    return "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )


# ------------------------------------------------------------

def normalizar_texto(texto):
    """
    Convierte un texto a un formato estándar:

    - Sin tildes
    - Sin espacios al principio y al final
    - En mayúsculas
    - Unifica saltos de línea
    - Elimina espacios repetidos
    """

    if texto is None:
        return ""

    texto = quitar_tildes(texto)

    texto = str(texto).strip().upper()

    # ----------------------------------------------------
    # Unificar saltos de línea y espacios repetidos
    # ----------------------------------------------------

    texto = " ".join(
        texto.split()
    )

    return texto


# ------------------------------------------------------------

def obtener_clave_alumno(nombre_completo):
    """
    Genera una clave única para identificar al alumno.

    Utiliza el nombre completo normalizado:
    - Sin tildes
    - Sin espacios al principio y al final
    - En mayúsculas

    El apellido NO se utiliza como identificador.
    """

    if nombre_completo is None:
        return ""

    return normalizar_texto(
        str(nombre_completo)
    )

# ------------------------------------------------------------

def buscar_fila_encabezado(df):
    """
    Busca automáticamente la fila donde comienza la tabla.
    """

    for fila in range(min(15, len(df))):

        valores = [
            normalizar_texto(x)
            for x in df.iloc[fila].fillna("")
        ]

        if normalizar_texto(COLUMNA_ALUMNO) in valores:
            return fila

    return None


# ------------------------------------------------------------

def convertir_nota(valor):
    """
    Convierte una nota al formato numérico.

    A -> 0

    Devuelve None si está vacía.
    """

    if pd.isna(valor):
        return None

    texto = str(valor).strip().upper()

    if texto == "":
        return None

    if texto == "A":
        return 0

    try:
        return float(texto)
    except:
        return None


# ------------------------------------------------------------

def es_nota_valida(valor):
    """
    Verifica si una nota es válida.
    """

    nota = convertir_nota(valor)

    if nota is None:
        return False

    return NOTA_MINIMA <= nota <= NOTA_MAXIMA

def obtener_apellido(nombre_completo):
    """
    Devuelve el primer apellido de un alumno.
    """

    if nombre_completo is None:
        return ""

    partes = normalizar_texto(nombre_completo).split()

    if len(partes) == 0:
        return ""

    return partes[0]

# ------------------------------------------------------------

def obtener_nota(fila, columna):
    """
    Obtiene una nota de una fila del DataFrame.

    Devuelve:
        - Número (0 a 10)
        - None si la columna no existe o la celda está vacía.

    Convierte automáticamente:
        A -> 0
    """

    if columna not in fila:
        return None

    return convertir_nota(fila[columna])


# ------------------------------------------------------------

def calcular_promedio(nota1, nota2):
    """
    Calcula el promedio de dos notas.

    Devuelve el resultado con un decimal.
    """

    if nota1 is None or nota2 is None:
        return None

    return round((nota1 + nota2) / 2, 1)


# ============================================================
# MÓDULO 1 - IMPORTACIONES
# ============================================================

import os
import sys
import unicodedata
import tkinter as tk
from tkinter import filedialog, messagebox

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border
from openpyxl.utils import get_column_letter

# Compatibilidad con archivos ODS
try:
    import odf
except ImportError:
    odf = None


# ============================================================
# MÓDULO 1 - CONFIGURACIÓN GENERAL
# ============================================================

VERSION = "4.0"

TITULO = f"Procesador de Cursadas V{VERSION}"


# ============================================================
# HOJAS DEL ARCHIVO
# ============================================================

HOJA_REPORTE = "REPORTE1"
HOJA_LIB = "LIB"
HOJA_REG2ET = "REG 2ET"
HOJA_INAS = "INAS"
HOJA_INAS2 = "INAS2"
HOJA_RCP = "RCP"
HOJA_PRN = "PRN"


HOJAS_OBLIGATORIAS = [
    HOJA_REPORTE,
    HOJA_LIB,
    HOJA_REG2ET,
    HOJA_INAS,
    HOJA_INAS2
]


# ============================================================
# COLUMNAS
# ============================================================

COLUMNA_ALUMNO = "ALUMNO"

COLUMNA_PROPUESTA = "PROPUESTA"

COLUMNAS_TP = [
    "TP1",
    "TP2",
    "TP3",
    "TP4"
]

COLUMNAS_PARCIALES = [
    "1P",
    "2P"
]

COLUMNAS_PROMEDIOS = [
    "PRM1",
    "PRM2"
]

COLUMNA_INAS = "INAS"


COLUMNAS_REPORTE = [
    COLUMNA_ALUMNO,
    "TP1",
    "TP2",
    "PRM1",
    "1P",
    "TP3",
    "TP4",
    "PRM2",
    "2P",
    COLUMNA_INAS
]


# ============================================================
# COLORES
# ============================================================

COLOR_VERDE = "92D050"
COLOR_AMARILLO = "FFD966"
COLOR_ROJO = "FF6666"

COLOR_ENCABEZADO = "D9EAD3"


# ============================================================
# VALIDACIONES
# ============================================================

NOTA_MINIMA = 0
NOTA_MAXIMA = 10

PROMEDIO_APROBACION = 4
PROMEDIO_PROMOCION = 7

NOTA_MINIMA_PROMOCION = 6

VALORES_ESPECIALES = [
    "A"
]


# ============================================================
# VARIABLES GLOBALES
# ============================================================

diferencias_1et = False

limite_faltas = None
archivo_salida = None

archivo_notas = None
archivo_inasistencias = None

workbook = None
worksheet = None

df_notas = None
df_reporte = None
df_inas1 = None
df_inas2 = None

alumnos_rcp = []

apellidos_repetidos = []

errores = []

lista_lib = []
inas_segunda_etapa = {}

# ============================================================
# FIN DEL MÓDULO 1
# ============================================================

print("=" * 55)
print(TITULO)
print("Módulo 1 cargado correctamente.")
print("=" * 55)

# ============================================================
# MÓDULO 2
# SELECCIÓN DE ARCHIVOS
# ============================================================

def seleccionar_archivo(titulo, tipos_archivo):
    """
    Abre una ventana para seleccionar un archivo.

    tipos_archivo:
        Lista de tuplas con la descripción y extensión.
    """

    ruta = filedialog.askopenfilename(
        title=titulo,
        filetypes=tipos_archivo
    )

    if not ruta:
        return None

    return ruta

# ------------------------------------------------------------

def seleccionar_archivos():

    global archivo_notas
    global archivo_inasistencias

    # ----------------------------------------------------
    # ARCHIVO DE NOTAS
    # ----------------------------------------------------

    actualizar_progreso(
        "Seleccionando archivo de notas...",
        5
    )

    root = tk.Tk()
    root.withdraw()

    archivo_notas = seleccionar_archivo(
        "Seleccione el archivo de NOTAS",
        [
            ("Archivos Excel y ODS", "*.xlsx *.ods"),
            ("Excel", "*.xlsx"),
            ("OpenDocument", "*.ods")
        ]
    )

    if archivo_notas is None:

        mostrar_advertencia(
            "Operación cancelada."
        )

        return False

    # ----------------------------------------------------
    # ARCHIVO DE INASISTENCIAS
    # ----------------------------------------------------

    actualizar_progreso(
        "Seleccionando archivo de inasistencias...",
        10
    )

    archivo_inasistencias = seleccionar_archivo(
        "Seleccione el archivo de INASISTENCIAS",
        [
            ("Archivo Excel", "*.xlsx")
        ]
    )

    if archivo_inasistencias is None:

        mostrar_advertencia(
            "Operación cancelada."
        )

        return False

    mostrar_info(
        "Archivos seleccionados correctamente."
    )

    actualizar_progreso(
        "Archivos seleccionados.",
        15
    )

    return True

# ============================================================
# MÓDULO 3
# VALIDACIÓN DE ARCHIVOS
# ============================================================

def validar_archivo_notas():

    global archivo_notas

    try:

        if archivo_notas.lower().endswith(".xlsx"):

            df = pd.read_excel(
                archivo_notas,
                header=None
            )

        else:

            df = pd.read_excel(
                archivo_notas,
                engine="odf",
                header=None
            )

    except Exception as e:

        mostrar_error(f"No fue posible abrir el archivo de notas.\n\n{e}")

        return False

    encontrado = False

    for fila in range(min(15, len(df))):

        for valor in df.iloc[fila]:

            if pd.isna(valor):
                continue

            if str(valor).strip().upper() == COLUMNA_PROPUESTA:

                encontrado = True
                break

        if encontrado:
            break

    if not encontrado:

        mostrar_error(
           "El archivo seleccionado no corresponde al archivo de NOTAS.\n\n"
            "No se encontró la columna PROPUESTA."
        )

        return False

    return True


# ------------------------------------------------------------

def validar_archivo_inasistencias():

    global archivo_inasistencias

    try:

        if archivo_inasistencias.lower().endswith(".xlsx"):

            libro = load_workbook(
                archivo_inasistencias,
                read_only=True
            )

            hojas = libro.sheetnames

            libro.close()

        elif archivo_inasistencias.lower().endswith(".ods"):

            libro = pd.ExcelFile(
                archivo_inasistencias,
                engine="odf"
            )

            hojas = libro.sheet_names

            libro.close()

        else:

            mostrar_error(
                "El archivo de inasistencias debe ser "
                "un archivo .xlsx o .ods."
            )

            return False

    except Exception as e:

        mostrar_error(
            f"No fue posible abrir el archivo de inasistencias.\n\n{e}"
        )

        return False

    faltantes = []

    for hoja in HOJAS_OBLIGATORIAS:

        if hoja not in hojas:

            faltantes.append(hoja)

    if faltantes:

        texto = "\n".join(faltantes)

        mostrar_error(
            "El archivo de inasistencias no es válido.\n\n"
            "Faltan las siguientes hojas:\n\n"
            f"{texto}"
        )

        return False

    return True


# ------------------------------------------------------------

def validar_cantidad_alumnos():

    """
    Se implementará en el Módulo 4,
    cuando ya tengamos ambos DataFrames cargados.
    """

    return True


# ------------------------------------------------------------

def validar_archivos():

    if not validar_archivo_notas():
        return False

    if not validar_archivo_inasistencias():
        return False

    actualizar_progreso(
        "Archivos validados correctamente.",
        45
    )

    return True

# ============================================================
# MÓDULO 4
# LECTURA DE DATOS
# ============================================================

def validar_coincidencia_alumnos():

    actualizar_progreso(
        "Verificando coincidencia de alumnos...",
        35
    )

    alumnos_notas = obtener_lista_alumnos(
        df_notas
    )

    alumnos_reporte = obtener_lista_alumnos(
        df_reporte
    )

    faltan_en_reporte = sorted(
        set(alumnos_notas.keys()) -
        set(alumnos_reporte.keys())
    )

    faltan_en_notas = sorted(
        set(alumnos_reporte.keys()) -
        set(alumnos_notas.keys())
    )

    # ----------------------------------------------------
    # SI COINCIDEN TODOS LOS ALUMNOS
    # ----------------------------------------------------

    if not faltan_en_reporte and not faltan_en_notas:

        return True

    # ----------------------------------------------------
    # CONSTRUIR MENSAJE DE ERROR
    # ----------------------------------------------------

    mensaje = ""

    if faltan_en_reporte:

        mensaje += (
            "Alumnos presentes en NOTAS y ausentes "
            "en REPORTE:\n\n"
        )

        for alumno in faltan_en_reporte:

            mensaje += (
                f"• {alumno}\n"
            )

    if faltan_en_notas:

        if mensaje != "":

            mensaje += (
                "\n-----------------------------\n\n"
            )

        mensaje += (
            "Alumnos presentes en REPORTE y ausentes "
            "en NOTAS:\n\n"
        )

        for alumno in faltan_en_notas:

            mensaje += (
                f"• {alumno}\n"
            )

    mostrar_error(
        mensaje
    )

    return False
    
def cargar_archivo_notas():

    global df_notas

    actualizar_progreso(
        "Leyendo archivo de notas...",
        50
    )

    try:

        if archivo_notas.lower().endswith(".xlsx"):

            bruto = pd.read_excel(
                archivo_notas,
                header=None
            )

        else:

            bruto = pd.read_excel(
                archivo_notas,
                engine="odf",
                header=None
            )

    except Exception as e:

        mostrar_error(
            f"No fue posible abrir el archivo de notas.\n\n{e}"
        )

        return False

    fila = buscar_fila_encabezado(bruto)

    if fila is None:

        mostrar_error(
            "No fue posible localizar el encabezado del archivo."
        )

        return False

    try:

        if archivo_notas.lower().endswith(".xlsx"):

            df_notas = pd.read_excel(
                archivo_notas,
                header=fila
            )

        else:

            df_notas = pd.read_excel(
                archivo_notas,
                engine="odf",
                header=fila
            )

    except Exception as e:

        mostrar_error(
            f"No fue posible leer los datos del archivo de notas.\n\n{e}"
        )

        return False

    # --------------------------------------------------------
    # NORMALIZAR COLUMNAS DEL ARCHIVO DE NOTAS
    # --------------------------------------------------------

    columnas = df_notas.columns.tolist()

    if len(columnas) < 7:

        mostrar_error(
            "El archivo de notas no contiene la cantidad "
            "de columnas esperada para la segunda etapa."
        )

        return False

    df_notas.columns = [
        "ALUMNO",
        "TP1",
        "TP2",
        "1P",
        "TP3",
        "TP4",
        "PROPUESTA"
    ]

    return True

def validar_notas_completas():
    """
    Verifica que todos los alumnos tengan notas válidas
    en TP1, TP2, 1P, TP3 y TP4.

    Se consideran inválidos:

        - Celdas vacías.
        - Valores que no puedan convertirse en nota.
        - Notas menores que 0.
        - Notas mayores que 10.

    La letra A se considera válida y equivale a 0.

    Si encuentra errores:
        - Muestra una lista con los alumnos afectados.
        - Indica qué nota tiene el problema.
        - Detiene el procesamiento.

    Devuelve:
        True  -> todas las notas son válidas.
        False -> hay notas inválidas o faltantes.
    """

    actualizar_progreso(
        "Verificando notas...",
        30
    )

    columnas_notas = [
        "TP1",
        "TP2",
        "1P",
        "TP3",
        "TP4"
    ]

    notas_invalidas = []

    # ----------------------------------------------------
    # RECORRER TODOS LOS ALUMNOS
    # ----------------------------------------------------

    for indice in df_notas.index:

        alumno = df_notas.at[
            indice,
            "ALUMNO"
        ]

        if pd.isna(alumno):
            continue

        alumno = str(alumno).strip()

        if alumno == "":
            continue

        # ------------------------------------------------
        # REVISAR CADA NOTA
        # ------------------------------------------------

        for columna in columnas_notas:

            valor = df_notas.at[
                indice,
                columna
            ]

            # --------------------------------------------
            # CELDA VACÍA
            # --------------------------------------------

            if pd.isna(valor):

                notas_invalidas.append(
                    f"• {alumno} — {columna}: VACÍA"
                )

                continue

            texto = str(valor).strip()

            if texto == "":

                notas_invalidas.append(
                    f"• {alumno} — {columna}: VACÍA"
                )

                continue

            # --------------------------------------------
            # NOTA INVÁLIDA
            # --------------------------------------------

            nota = convertir_nota(valor)

            if nota is None:

                notas_invalidas.append(
                    f"• {alumno} — {columna}: "
                    f"'{texto}' no es una nota válida"
                )

                continue

            # --------------------------------------------
            # NOTA FUERA DEL RANGO
            # --------------------------------------------

            if (
                nota < NOTA_MINIMA
                or
                nota > NOTA_MAXIMA
            ):

                notas_invalidas.append(
                    f"• {alumno} — {columna}: "
                    f"{nota} fuera del rango permitido "
                    f"({NOTA_MINIMA} a {NOTA_MAXIMA})"
                )

                continue

            # --------------------------------------------
            # NOTA CON DECIMALES
            # --------------------------------------------

            if not nota.is_integer():

                notas_invalidas.append(
                    f"• {alumno} — {columna}: "
                    f"{nota} no es una nota entera"
                )

                continue

    # ----------------------------------------------------
    # TODAS LAS NOTAS SON VÁLIDAS
    # ----------------------------------------------------

    if not notas_invalidas:
        return True

    # ----------------------------------------------------
    # HAY NOTAS INVÁLIDAS
    # ----------------------------------------------------

    mensaje = (
        "NO SE PUEDE CONTINUAR.\n\n"
        "Se encontraron notas faltantes o inválidas:\n\n"
        + "\n".join(notas_invalidas)
        + "\n\n"
        "Corrija las notas antes de volver a ejecutar "
        "el programa."
    )

    mostrar_error(mensaje)

    return False

def cargar_archivo_inasistencias():

    global df_reporte
    global df_inas
    global df_inas2

    actualizar_progreso(
        "Leyendo archivo de inasistencias...",
        60
    )

    try:

        # ----------------------------------------------------
        # REPORTE
        # ----------------------------------------------------

        bruto_reporte = pd.read_excel(
            archivo_inasistencias,
            sheet_name=HOJA_REPORTE,
            header=None
        )

        fila_reporte = buscar_fila_encabezado(
            bruto_reporte
        )

        if fila_reporte is None:

            mostrar_error(
                "No fue posible localizar el encabezado "
                "de la hoja REPORTE."
            )

            return False

        df_reporte = pd.read_excel(
            archivo_inasistencias,
            sheet_name=HOJA_REPORTE,
            header=fila_reporte
        )

        # ----------------------------------------------------
        # INAS
        # ----------------------------------------------------

        df_inas = pd.read_excel(
            archivo_inasistencias,
            sheet_name=HOJA_INAS
        )

        # ----------------------------------------------------
        # INAS2
        # ----------------------------------------------------

        df_inas2 = pd.read_excel(
            archivo_inasistencias,
            sheet_name=HOJA_INAS2
        )

    except Exception as e:

        mostrar_error(
            "No fue posible abrir el archivo "
            "de inasistencias.\n\n"
            f"{e}"
        )

        return False
        # Normalizar columnas de INAS e INAS2
    df_inas = df_inas.rename(
        columns={"Apellido y nombre": "ALUMNO"}
        )

    df_inas2 = df_inas2.rename(
        columns={"Apellido y nombre": "ALUMNO"}
        )

    return True

def leer_datos():
    """
    Lee ambos archivos y verifica que puedan procesarse.
    """

    if not cargar_archivo_notas():
        return False

    if not validar_notas_completas():
        return False

    if not cargar_archivo_inasistencias():
        return False

    # ----------------------------------------------------
    # REGISTRAR APELLIDOS REPETIDOS
    # ----------------------------------------------------

    global apellidos_repetidos

    apellidos_repetidos = []

    obtener_lista_alumnos(df_notas)
    obtener_lista_alumnos(df_reporte)
    obtener_lista_alumnos(df_inas)
    obtener_lista_alumnos(df_inas2)

    # ----------------------------------------------------
    # VERIFICAR COINCIDENCIA DE ALUMNOS
    # ----------------------------------------------------

    if not validar_coincidencia_alumnos():
        return False

    actualizar_progreso(
        "Datos cargados correctamente.",
        65
    )

    return True

# ============================================================
# MÓDULO 5
# PROCESAMIENTO DE SEGUNDA ETAPA
# ============================================================

def calcular_prm2():
    """
    Calcula PRM2 como el promedio de TP3 y TP4.
    El resultado se redondea a un decimal.
    """

    global df_notas

    actualizar_progreso(
        "Calculando PRM2...",
        45
    )

    for indice in df_notas.index:

        tp3 = obtener_nota(
            df_notas.loc[indice],
            "TP3"
            )

        tp4 = obtener_nota(
            df_notas.loc[indice],
            "TP4"
            )

        prm2 = round((tp3 + tp4) / 2, 1)

        df_notas.at[indice, "PRM2"] = prm2

    return True


def determinar_habilitados_segundo_parcial():
    """
    Determina qué alumnos pueden rendir el segundo parcial.

    PRM2 >= 4:
        Puede rendir.

    PRM2 < 4:
        No puede rendir y pasa a LIB.
    """

    global df_notas
    global lista_lib

    actualizar_progreso(
        "Verificando alumnos habilitados para el segundo parcial...",
        50
    )

    lista_lib = []

    for indice in df_notas.index:

        alumno = df_notas.at[indice, "ALUMNO"]

        prm2 = obtener_nota(
            df_notas.loc[indice],
            "PRM2"
            )

        if prm2 < 4:

            lista_lib.append({
                "alumno": alumno,
                "fila": indice,
                "prm2": prm2
            })

    return True


def calcular_inasistencias_segunda_etapa():
    """
    Calcula las inasistencias correspondientes
    únicamente a la segunda etapa.

    INAS segunda etapa =
        INAS2 acumuladas - INAS acumuladas.

    El alumno se identifica mediante su nombre completo
    normalizado.

    Si el resultado es negativo, se considera 0.
    """

    global df_reporte
    global df_inas
    global df_inas2
    global inas_segunda_etapa

    actualizar_progreso(
        "Calculando inasistencias de la segunda etapa...",
        55
    )

    # ----------------------------------------------------
    # Verificar que existan las columnas necesarias
    # ----------------------------------------------------

    columna_inas = "Inasistencias acumuladas"

    if columna_inas not in df_inas.columns:

        mostrar_error(
            "La hoja INAS no contiene la columna "
            "'Inasistencias acumuladas'."
        )

        return False

    if columna_inas not in df_inas2.columns:

        mostrar_error(
            "La hoja INAS2 no contiene la columna "
            "'Inasistencias acumuladas'."
        )

        return False

    # ----------------------------------------------------
    # Inicializar diccionario
    # ----------------------------------------------------

    inas_segunda_etapa = {}

    # ----------------------------------------------------
    # Obtener alumnos de cada hoja
    # ----------------------------------------------------

    alumnos_reporte = obtener_lista_alumnos(
        df_reporte
    )

    alumnos_inas = obtener_lista_alumnos(
        df_inas
    )

    alumnos_inas2 = obtener_lista_alumnos(
        df_inas2
    )

    # ----------------------------------------------------
    # Recorrer todos los alumnos del REPORTE
    # ----------------------------------------------------

    for clave in alumnos_reporte.keys():

        valor_inas = 0
        valor_inas2 = 0

        # =================================================
        # INAS PRIMERA ETAPA
        # =================================================

        if clave in alumnos_inas:

            fila_inas = alumnos_inas[clave]["fila"]

            valor = df_inas.loc[
                fila_inas,
                columna_inas
            ]

            if pd.notna(valor):

                try:

                    valor_inas = float(valor)

                except (ValueError, TypeError):

                    valor_inas = 0

        # =================================================
        # INAS2 ACUMULADAS
        # =================================================

        if clave in alumnos_inas2:

            fila_inas2 = alumnos_inas2[clave]["fila"]

            valor = df_inas2.loc[
                fila_inas2,
                columna_inas
            ]

            if pd.notna(valor):

                try:

                    valor_inas2 = float(valor)

                except (ValueError, TypeError):

                    valor_inas2 = 0

        # =================================================
        # CALCULAR DIFERENCIA
        # =================================================

        resultado = (
            valor_inas2
            -
            valor_inas
        )

        # -------------------------------------------------
        # Evitar valores negativos
        # -------------------------------------------------

        if resultado < 0:

            resultado = 0

        # -------------------------------------------------
        # Guardar resultado
        # -------------------------------------------------

        inas_segunda_etapa[clave] = resultado

    return True


def solicitar_limite_faltas():
    """
    Solicita al usuario el límite de faltas permitido
    para la materia.
    """

    global limite_faltas

    while True:

        limite = simpledialog.askinteger(
            TITULO,
            "¿Cuál es el límite de faltas para esta materia?",
            minvalue=0
        )

        if limite is None:

            mostrar_error(
                "No se indicó el límite de faltas.\n\n"
                "La operación será cancelada."
            )

            return False

        limite_faltas = limite

        return True

def preparar_datos_segunda_etapa():
    """
    Ejecuta todo el procesamiento correspondiente
    a la segunda etapa.
    """

    actualizar_progreso(
        "Procesando segunda etapa...",
        40
    )

    if not solicitar_limite_faltas():
        return False

    if not calcular_prm2():
        return False

    if not determinar_habilitados_segundo_parcial():
        return False

    if not calcular_inasistencias_segunda_etapa():
        return False

    actualizar_progreso(
        "Procesamiento de segunda etapa completado.",
        65
    )

    return True


# ============================================================
# MÓDULO 6
# GENERACIÓN DEL ARCHIVO FINAL
# ============================================================

import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font



def obtener_datos_notas():
    """
    Convierte el DataFrame de notas en un diccionario
    indexado por el nombre completo normalizado del alumno.

    El nombre completo se utiliza como identificador
    para evitar conflictos entre alumnos con el mismo apellido.
    """

    datos = {}

    for indice in df_notas.index:

        alumno = df_notas.at[
            indice,
            "ALUMNO"
        ]

        if pd.isna(alumno):
            continue

        alumno = str(alumno).strip()

        if alumno == "":
            continue

        clave = obtener_clave_alumno(
            alumno
        )

        datos[clave] = {

            "alumno": alumno,

            "TP1": obtener_nota(
                df_notas.loc[indice],
                "TP1"
            ),

            "TP2": obtener_nota(
                df_notas.loc[indice],
                "TP2"
            ),

            "1P": obtener_nota(
                df_notas.loc[indice],
                "1P"
            ),

            "TP3": obtener_nota(
                df_notas.loc[indice],
                "TP3"
            ),

            "TP4": obtener_nota(
                df_notas.loc[indice],
                "TP4"
            )
        }

    return datos


def crear_archivo_salida():
    """
    Crea una copia del archivo de inasistencias
    que será utilizada como archivo final.

    De esta manera se conserva el formato original.
    """

    actualizar_progreso(
        "Creando archivo de salida...",
        70
    )

    carpeta = os.path.dirname(
        archivo_inasistencias
    )

    nombre = os.path.splitext(
        os.path.basename(archivo_inasistencias)
    )[0]

    archivo_salida = os.path.join(
        carpeta,
        nombre + "_procesado.xlsx"
    )

    shutil.copy2(
        archivo_inasistencias,
        archivo_salida
    )

    return archivo_salida


def eliminar_hojas_no_necesarias(wb):
    """
    Deja únicamente las cuatro hojas necesarias:

        REPORTE
        LIB
        INAS
        INAS2

    Si el archivo original contiene REPORTE1,
    la renombra a REPORTE.
    """

    # ----------------------------------------------------
    # Convertir REPORTE1 en REPORTE
    # ----------------------------------------------------

    if "REPORTE1" in wb.sheetnames:

        if "REPORTE" in wb.sheetnames:
            del wb["REPORTE"]

        wb["REPORTE1"].title = "REPORTE"

    # ----------------------------------------------------
    # Hojas necesarias en el archivo final
    # ----------------------------------------------------

    hojas_necesarias = {
        "REPORTE",
        "LIB",
        "INAS",
        "INAS2"
    }

    # ----------------------------------------------------
    # Eliminar todas las demás hojas
    # ----------------------------------------------------

    for nombre in list(wb.sheetnames):

        if nombre not in hojas_necesarias:

            del wb[nombre]


def obtener_fila_encabezado(ws):
    """
    Busca la fila que contiene ALUMNO.

    Se utiliza para no depender rígidamente
    de una posición determinada.
    """

    for fila in range(1, ws.max_row + 1):

        for columna in range(1, ws.max_column + 1):

            valor = ws.cell(
                fila,
                columna
            ).value

            if valor is None:
                continue

            texto = str(valor).strip().upper()

            if texto == "ALUMNO":
                return fila

    return None


def obtener_columnas_reporte(ws):
    """
    Obtiene las columnas de la hoja REPORTE
    según sus encabezados.
    """

    fila_encabezado = obtener_fila_encabezado(ws)

    if fila_encabezado is None:

        mostrar_error(
            "No se encontró la fila de encabezados "
            "en la hoja REPORTE."
        )

        return None

    columnas = {}

    for columna in range(
        1,
        ws.max_column + 1
    ):

        valor = ws.cell(
            fila_encabezado,
            columna
        ).value

        if valor is None:
            continue

        texto = str(
            valor
        ).strip().upper()

        if texto == "ALUMNO":
            columnas["ALUMNO"] = columna

        elif texto == "TP1":
            columnas["TP1"] = columna

        elif texto == "TP2":
            columnas["TP2"] = columna

        elif texto == "PRM1":
            columnas["PRM1"] = columna

        elif texto == "1P":
            columnas["1P"] = columna

        elif texto == "TP3":
            columnas["TP3"] = columna

        elif texto == "TP4":
            columnas["TP4"] = columna

        elif texto == "PRM2":
            columnas["PRM2"] = columna

        elif texto == "INAS":
            columnas["INAS"] = columna

        elif texto == "OBSERVACIONES":
            columnas["OBSERVACION"] = columna

    return fila_encabezado, columnas


def preparar_reporte(ws):
    """
    Prepara la hoja REPORTE para recibir
    la información de la segunda etapa.
    """

    resultado = obtener_columnas_reporte(ws)

    if resultado is None:
        return False

    fila_encabezado, columnas = resultado

    actualizar_progreso(
        "Preparando hoja REPORTE...",
        75
    )

    columnas_necesarias = [
        "ALUMNO",
        "TP1",
        "TP2",
        "PRM1",
        "1P",
        "TP3",
        "TP4",
        "PRM2",
        "INAS",
        "OBSERVACION"
    ]

    ultima_columna = ws.max_column

    for nombre in columnas_necesarias:

        if nombre not in columnas:

            ultima_columna += 1

            ws.cell(
                fila_encabezado,
                ultima_columna
            ).value = nombre

            columnas[nombre] = ultima_columna

    return fila_encabezado, columnas

def preparar_reporte_final(ws):
    """
    Prepara la hoja REPORTE desde cero.

    Estructura definitiva:

    ALUMNO
    TP1
    TP2
    PRM1
    1P
    TP3
    TP4
    PRM2
    INAS
    OBSERVACIONES

    Se elimina el formato heredado y se aplica:
    - Fondo blanco
    - Encabezado gris claro
    - Texto negro
    """

    # ----------------------------------------------------
    # Buscar encabezado actual
    # ----------------------------------------------------

    fila_origen = None

    for fila in range(
        1,
        ws.max_row + 1
    ):

        valores = []

        for columna in range(
            1,
            ws.max_column + 1
        ):

            valor = ws.cell(
                fila,
                columna
            ).value

            if valor is not None:

                valores.append(
                    normalizar_texto(
                        str(valor)
                    )
                )

        if "ALUMNO" in valores:

            fila_origen = fila

            break

    # ----------------------------------------------------
    # Verificar que encontramos el encabezado
    # ----------------------------------------------------

    if fila_origen is None:

        mostrar_error(
            "No fue posible localizar el encabezado "
            "ALUMNO en la hoja REPORTE."
        )

        return False

    # ----------------------------------------------------
    # Buscar columna ALUMNO
    # ----------------------------------------------------

    columna_alumno = None

    for columna in range(
        1,
        ws.max_column + 1
    ):

        valor = ws.cell(
            fila_origen,
            columna
        ).value

        if valor is None:
            continue

        if normalizar_texto(valor) == "ALUMNO":

            columna_alumno = columna

            break

    if columna_alumno is None:

        mostrar_error(
            "No fue posible localizar la columna "
            "ALUMNO en la hoja REPORTE."
        )

        return False

    # ----------------------------------------------------
    # Obtener alumnos actuales
    # ----------------------------------------------------

    alumnos = []

    for fila in range(
        fila_origen + 1,
        ws.max_row + 1
    ):

        valor = ws.cell(
            fila,
            columna_alumno
        ).value

        if valor is None:
            continue

        alumno = str(
            valor
        ).strip()

        if alumno == "":
            continue

        alumnos.append(
            alumno
        )

    # ----------------------------------------------------
    # Limpiar completamente la hoja
    # ----------------------------------------------------

    for fila in range(
        1,
        ws.max_row + 1
    ):

        for columna in range(
            1,
            ws.max_column + 1
        ):

            celda = ws.cell(
                fila,
                columna
            )

            celda.value = None

            celda.fill = PatternFill(
                fill_type=None
            )

            celda.font = Font(
                color="000000"
            )

            celda.border = Border()

            celda.alignment = Alignment(
                vertical="center"
            )

    # ----------------------------------------------------
    # Encabezados definitivos
    # ----------------------------------------------------

    encabezados = [
        "ALUMNO",
        "TP1",
        "TP2",
        "PRM1",
        "1P",
        "TP3",
        "TP4",
        "PRM2",
        "INAS",
        "OBSERVACIONES"
    ]

    # ----------------------------------------------------
    # Formato del encabezado
    # ----------------------------------------------------

    encabezado_fill = PatternFill(
        fill_type="solid",
        fgColor="D9E1F2"
    )

    encabezado_font = Font(
        bold=True,
        color="000000"
    )

    for columna, nombre in enumerate(
        encabezados,
        start=1
    ):

        celda = ws.cell(
            1,
            columna
        )

        celda.value = nombre

        celda.fill = encabezado_fill

        celda.font = encabezado_font

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # ----------------------------------------------------
    # Escribir alumnos
    # ----------------------------------------------------

    for fila, alumno in enumerate(
        alumnos,
        start=2
    ):

        ws.cell(
            fila,
            1
        ).value = alumno

    # ----------------------------------------------------
    # Formato general
    # ----------------------------------------------------

    fondo_blanco = PatternFill(
        fill_type="solid",
        fgColor="FFFFFF"
    )

    fuente_negra = Font(
        color="000000"
    )

    for fila in range(
        2,
        len(alumnos) + 2
    ):

        for columna in range(
            1,
            11
        ):

            celda = ws.cell(
                fila,
                columna
            )

            celda.fill = fondo_blanco

            celda.font = fuente_negra

            celda.alignment = Alignment(
                vertical="center"
            )

    # ----------------------------------------------------
    # Anchos de columnas
    # ----------------------------------------------------

    ws.column_dimensions["A"].width = 35

    for columna in [
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I"
    ]:

        ws.column_dimensions[
            columna
        ].width = 12

    ws.column_dimensions["J"].width = 38

    # ----------------------------------------------------
    # Inmovilizar encabezado
    # ----------------------------------------------------

    ws.freeze_panes = "A2"

    return True

def preparar_reporte_salida(ws):
    """
    Prepara la hoja REPORTE para la segunda etapa.

    Conserva el formato original del archivo descargado.
    Elimina las columnas que no necesitamos y agrega
    las nuevas columnas manteniendo el orden definitivo.
    """

    actualizar_progreso(
        "Preparando hoja REPORTE...",
        75
    )

    # ----------------------------------------------------
    # Buscar encabezado
    # ----------------------------------------------------

    fila_encabezado = obtener_fila_encabezado(ws)

    if fila_encabezado is None:

        mostrar_error(
            "No se pudo encontrar el encabezado "
            "de la hoja REPORTE."
        )

        return False

    # ----------------------------------------------------
    # Leer encabezados actuales
    # ----------------------------------------------------

    encabezados_actuales = {}

    for columna in range(
        1,
        ws.max_column + 1
    ):

        valor = ws.cell(
            fila_encabezado,
            columna
        ).value

        if valor is None:
            continue

        texto = str(
            valor
        ).strip().upper()

        encabezados_actuales[texto] = columna

    # ----------------------------------------------------
    # Verificar ALUMNO
    # ----------------------------------------------------

    if "ALUMNO" not in encabezados_actuales:

        mostrar_error(
            "No se pudo encontrar la columna "
            "'ALUMNO' en la hoja REPORTE."
        )

        return False

    # ----------------------------------------------------
    # Guardar los alumnos antes de modificar columnas
    # ----------------------------------------------------

    alumnos = []

    columna_alumno = encabezados_actuales["ALUMNO"]

    for fila in range(
        fila_encabezado + 1,
        ws.max_row + 1
    ):

        valor = ws.cell(
            fila,
            columna_alumno
        ).value

        if valor is None:
            continue

        alumno = str(
            valor
        ).strip()

        if alumno == "":
            continue

        alumnos.append(
            alumno
        )

    # ----------------------------------------------------
    # Eliminar columnas posteriores a ALUMNO
    #
    # Esto permite reconstruir únicamente las columnas
    # que necesitamos, conservando la primera columna
    # y su formato original.
    # ----------------------------------------------------

    if ws.max_column > 1:

        ws.delete_cols(
            2,
            ws.max_column - 1
        )

    # ----------------------------------------------------
    # Encabezados definitivos
    # ----------------------------------------------------

    encabezados = [
        "ALUMNO",
        "TP1",
        "TP2",
        "PRM1",
        "1P",
        "TP3",
        "TP4",
        "PRM2",
        "INAS",
        "OBSERVACIONES"
    ]

    # ----------------------------------------------------
    # Escribir encabezados
    # ----------------------------------------------------

    for columna, nombre in enumerate(
        encabezados,
        start=1
    ):

        celda = ws.cell(
            fila_encabezado,
            columna
        )

        celda.value = nombre

    # ----------------------------------------------------
    # Restaurar alumnos
    # ----------------------------------------------------

    for fila, alumno in enumerate(
        alumnos,
        start=fila_encabezado + 1
    ):

        ws.cell(
            fila,
            1
        ).value = alumno

    # ----------------------------------------------------
    # Limpiar las nuevas columnas
    # ----------------------------------------------------

    for fila in range(
        fila_encabezado + 1,
        ws.max_row + 1
    ):

        for columna in range(
            2,
            11
        ):

            ws.cell(
                fila,
                columna
            ).value = None

    # ----------------------------------------------------
    # Anchos
    # ----------------------------------------------------

    ws.column_dimensions["A"].width = 35

    for columna in [
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I"
    ]:

        ws.column_dimensions[
            columna
        ].width = 12

    ws.column_dimensions[
        "J"
    ].width = 38

    return True


def escribir_notas_reporte(ws):
    """
    Escribe las notas del archivo de notas
    en la hoja REPORTE.
    """
    if not preparar_reporte_final(ws):
        return False

    datos_notas = obtener_datos_notas()

    if datos_notas is None:
        return False

    actualizar_progreso(
        "Copiando notas al REPORTE...",
        80
        )

    # Columnas definitivas
    columnas = {
        "ALUMNO": 1,
        "TP1": 2,
        "TP2": 3,
        "PRM1": 4,
        "1P": 5,
        "TP3": 6,
        "TP4": 7,
        "PRM2": 8,
        "INAS": 9,
        "OBSERVACION": 10
        }

    for fila in range(
        2,
        ws.max_row + 1
        ):

        valor_alumno = ws.cell(
            fila,
            columnas["ALUMNO"]
        ).value

        if valor_alumno is None:
            continue

        alumno = str(
            valor_alumno
        ).strip()

        if alumno == "":
            continue

        clave = obtener_clave_alumno(
            alumno
            )
        if clave not in datos_notas:
            continue

        datos = datos_notas[clave]

        ws.cell(
            fila,
            columnas["TP1"]
        ).value = datos["TP1"]

        ws.cell(
            fila,
            columnas["TP2"]
        ).value = datos["TP2"]

        ws.cell(
            fila,
            columnas["1P"]
        ).value = datos["1P"]

        ws.cell(
            fila,
            columnas["TP3"]
        ).value = datos["TP3"]

        ws.cell(
            fila,
            columnas["TP4"]
        ).value = datos["TP4"]

        prm1 = round(
            (
                datos["TP1"]
                +
                datos["TP2"]
            ) / 2,
            1
        )

        prm2 = round(
            (
                datos["TP3"]
                +
                datos["TP4"]
            ) / 2,
            1
        )

        ws.cell(
            fila,
            columnas["PRM1"]
        ).value = prm1

        ws.cell(
            fila,
            columnas["PRM2"]
        ).value = prm2

    return True

def escribir_inasistencias_reporte(ws):
    """
    Escribe en REPORTE las inasistencias
    correspondientes únicamente a la segunda etapa.
    
    INAS = INAS2 - INAS primera etapa.
    """

    actualizar_progreso(
        "Copiando inasistencias...",
        85
        )
    columnas = {
        "ALUMNO": 1,
        "TP1": 2,
        "TP2": 3,
        "PRM1": 4,
        "1P": 5,
        "TP3": 6,
        "TP4": 7,
        "PRM2": 8,
        "INAS": 9,
        "OBSERVACION": 10
        }

    for fila in range(
        2,
        ws.max_row + 1
        ):

        valor_alumno = ws.cell(
            fila,
            columnas["ALUMNO"]
        ).value

        if valor_alumno is None:
            continue

        alumno = str(
            valor_alumno
        ).strip()

        if alumno == "":
            continue

        clave = obtener_clave_alumno(
            alumno
            )

        if clave in inas_segunda_etapa:
            ws.cell(
                fila,
                columnas["INAS"]
            ).value = inas_segunda_etapa[
            clave
            ]

        else:

            ws.cell(
                fila,
                columnas["INAS"]
            ).value = 0

    return True

def validar_diferencias_1et(wb):
    """
    Compara TP1 y TP2 entre el archivo de inasistencias
    y el archivo de notas.

    Genera la hoja 1ET con el resultado de la comparación.

    Devuelve:
        True  -> hay diferencias
        False -> no hay diferencias
    """

    global diferencias_1et

    diferencias_1et = False

    # ----------------------------------------------------
    # Eliminar 1ET si ya existiera
    # ----------------------------------------------------

    if "1ET" in wb.sheetnames:
        del wb["1ET"]

    ws_1et = wb.create_sheet("1ET")

    # ----------------------------------------------------
    # Encabezado
    # ----------------------------------------------------

    ws_1et["A1"] = "CONTROL DE PRIMERA ETAPA"

    ws_1et["A1"].font = Font(
        bold=True
    )

    # ----------------------------------------------------
    # Obtener columnas del REPORTE
    # ----------------------------------------------------

    resultado = obtener_columnas_reporte(
        wb["REPORTE"]
    )

    if resultado is None:
        return False

    fila_encabezado, columnas = resultado

    # ----------------------------------------------------
    # Obtener datos del archivo de notas
    # ----------------------------------------------------

    datos_notas = obtener_datos_notas()

    if datos_notas is None:
        return False

    # ----------------------------------------------------
    # Encabezados del detalle
    # ----------------------------------------------------

    ws_1et["A3"] = "ALUMNO"
    ws_1et["B3"] = "TP1 REPORTE"
    ws_1et["C3"] = "TP1 NOTAS"
    ws_1et["D3"] = "TP2 REPORTE"
    ws_1et["E3"] = "TP2 NOTAS"

    encabezado_fill = PatternFill(
        fill_type="solid",
        fgColor="D9D9D9"
    )

    encabezado_font = Font(
        bold=True,
        color="000000"
    )

    for columna in range(1, 6):

        celda = ws_1et.cell(
            3,
            columna
        )

        celda.fill = encabezado_fill
        celda.font = encabezado_font

    fila_salida = 4

    # ----------------------------------------------------
    # Recorrer alumnos del REPORTE
    # ----------------------------------------------------

    ws_reporte = wb["REPORTE"]

    for fila in range(
        fila_encabezado + 1,
        ws_reporte.max_row + 1
    ):

        valor_alumno = ws_reporte.cell(
            fila,
            columnas["ALUMNO"]
        ).value

        if valor_alumno is None:
            continue

        alumno = str(
            valor_alumno
        ).strip()

        if alumno == "":
            continue

        clave = obtener_clave_alumno(
            alumno
        )

        if clave not in datos_notas:
            continue

        datos = datos_notas[clave]

        # ------------------------------------------------
        # Valores del REPORTE
        # ------------------------------------------------

        tp1_reporte = convertir_nota(
            ws_reporte.cell(
                fila,
                columnas["TP1"]
            ).value
        )

        tp2_reporte = convertir_nota(
            ws_reporte.cell(
                fila,
                columnas["TP2"]
            ).value
        )

        # ------------------------------------------------
        # Valores del archivo de notas
        # ------------------------------------------------

        tp1_notas = convertir_nota(
            datos["TP1"]
        )

        tp2_notas = convertir_nota(
            datos["TP2"]
        )

        diferencia_tp1 = (
            tp1_reporte != tp1_notas
        )

        diferencia_tp2 = (
            tp2_reporte != tp2_notas
        )

        # ------------------------------------------------
        # Si hay alguna diferencia
        # ------------------------------------------------

        if diferencia_tp1 or diferencia_tp2:

            diferencias_1et = True

            ws_1et.cell(
                fila_salida,
                1
            ).value = alumno

            if diferencia_tp1:

                ws_1et.cell(
                    fila_salida,
                    2
                ).value = tp1_reporte

                ws_1et.cell(
                    fila_salida,
                    3
                ).value = tp1_notas

            if diferencia_tp2:

                ws_1et.cell(
                    fila_salida,
                    4
                ).value = tp2_reporte

                ws_1et.cell(
                    fila_salida,
                    5
                ).value = tp2_notas

            fila_salida += 1

    # ----------------------------------------------------
    # Resultado si no hay diferencias
    # ----------------------------------------------------

    if not diferencias_1et:

        ws_1et["A3"] = "SIN DIFERENCIA CON 1ET"

        ws_1et["A3"].font = Font(
            bold=True
        )

        ws_1et.column_dimensions["A"].width = 30

        return False

    # ----------------------------------------------------
    # Formato de la hoja
    # ----------------------------------------------------

    ws_1et.column_dimensions["A"].width = 35
    ws_1et.column_dimensions["B"].width = 15
    ws_1et.column_dimensions["C"].width = 15
    ws_1et.column_dimensions["D"].width = 15
    ws_1et.column_dimensions["E"].width = 15

    ws_1et.freeze_panes = "A4"

    return True

def marcar_libres_en_reporte(ws):
    """
    Marca en rojo en REPORTE a los alumnos que quedaron LIB.

    Los alumnos libres tienen:
        - Fondo rojo en toda la fila.
        - Fuente negra.

    Los demás alumnos conservan:
        - Fondo blanco.
        - Fuente negra.

    El alumno se identifica mediante su nombre completo
    normalizado, no mediante el apellido.
    """

    actualizar_progreso(
        "Marcando alumnos libres en REPORTE...",
        93
    )

    # ----------------------------------------------------
    # Buscar encabezado
    # ----------------------------------------------------

    fila_encabezado = obtener_fila_encabezado(ws)

    if fila_encabezado is None:

        mostrar_error(
            "No se pudo encontrar el encabezado "
            "de la hoja REPORTE."
        )

        return False

    # ----------------------------------------------------
    # Buscar columna ALUMNO
    # ----------------------------------------------------

    columna_alumno = None

    for columna in range(
        1,
        ws.max_column + 1
    ):

        valor = ws.cell(
            fila_encabezado,
            columna
        ).value

        if valor is None:
            continue

        if str(
            valor
        ).strip().upper() == "ALUMNO":

            columna_alumno = columna
            break

    if columna_alumno is None:

        mostrar_error(
            "No se encontró la columna "
            "'ALUMNO' en REPORTE."
        )

        return False

    # ----------------------------------------------------
    # Fondo blanco y fuente negra para todos
    # ----------------------------------------------------

    fondo_blanco = PatternFill(
        fill_type="solid",
        fgColor="FFFFFF"
    )

    fuente_negra = Font(
        color="000000"
    )

    for fila in range(
        fila_encabezado + 1,
        ws.max_row + 1
    ):

        for columna in range(
            1,
            ws.max_column + 1
        ):

            celda = ws.cell(
                fila,
                columna
            )

            celda.fill = fondo_blanco
            celda.font = fuente_negra

    # ----------------------------------------------------
    # Crear conjunto de alumnos libres
    # ----------------------------------------------------

    alumnos_libres = set()

    for alumno in lista_lib:

        nombre = alumno.get(
            "alumno"
        )

        if not nombre:
            continue

        clave = obtener_clave_alumno(
            str(nombre)
        )

        alumnos_libres.add(
            clave
        )

    # ----------------------------------------------------
    # Fondo rojo para alumnos LIB
    # ----------------------------------------------------

    fondo_rojo = PatternFill(
        fill_type="solid",
        fgColor="FF0000"
    )

    # ----------------------------------------------------
    # Buscar alumnos en REPORTE
    # ----------------------------------------------------

    for fila in range(
        fila_encabezado + 1,
        ws.max_row + 1
    ):

        valor = ws.cell(
            fila,
            columna_alumno
        ).value

        if valor is None:
            continue

        alumno = str(
            valor
        ).strip()

        if alumno == "":
            continue

        clave = obtener_clave_alumno(
            alumno
        )

        if clave in alumnos_libres:

            for columna in range(
                1,
                ws.max_column + 1
            ):

                celda = ws.cell(
                    fila,
                    columna
                )

                celda.fill = fondo_rojo
                celda.font = fuente_negra

    return True


def marcar_alumnos_excedidos_en_faltas(ws):
    """
    Aplica el formato final de las filas de REPORTE.

    ALUMNOS LIBRES:
        - Fondo rojo en toda la fila.
        - Fuente negra.

    ALUMNOS REGULARES EXCEDIDOS EN FALTAS:
        - Fondo blanco.
        - Fuente roja.
        - Observación:
          ALUMNO REGULAR EXCEDIDO EN FALTAS

    RESTO:
        - Fondo blanco.
        - Fuente negra.
        - Sin observación.

    El alumno se identifica mediante su nombre completo
    normalizado, no mediante el apellido.
    """

    actualizar_progreso(
        "Aplicando formato final al REPORTE...",
        88
    )

    columnas = {
        "ALUMNO": 1,
        "TP1": 2,
        "TP2": 3,
        "PRM1": 4,
        "1P": 5,
        "TP3": 6,
        "TP4": 7,
        "PRM2": 8,
        "INAS": 9,
        "OBSERVACION": 10
    }

    # ----------------------------------------------------
    # Colores
    # ----------------------------------------------------

    relleno_blanco = PatternFill(
        fill_type="solid",
        fgColor="FFFFFF"
    )

    relleno_rojo = PatternFill(
        fill_type="solid",
        fgColor="FF0000"
    )

    fuente_negra = Font(
        color="000000"
    )

    fuente_roja = Font(
        color="FF0000"
    )

    # ----------------------------------------------------
    # Crear conjunto de alumnos libres
    # ----------------------------------------------------

    alumnos_libres = set()

    for alumno in lista_lib:

        nombre = alumno.get(
            "alumno"
        )

        if nombre:

            clave = obtener_clave_alumno(
                str(nombre).strip()
            )

            alumnos_libres.add(
                clave
            )

    # ----------------------------------------------------
    # Recorrer REPORTE
    # ----------------------------------------------------

    for fila in range(
        2,
        ws.max_row + 1
    ):

        valor_alumno = ws.cell(
            fila,
            columnas["ALUMNO"]
        ).value

        if valor_alumno is None:
            continue

        alumno = str(
            valor_alumno
        ).strip()

        if alumno == "":
            continue

        clave = obtener_clave_alumno(
            alumno
        )

        # =================================================
        # ALUMNO LIBRE
        # =================================================

        if clave in alumnos_libres:

            ws.cell(
                fila,
                columnas["OBSERVACION"]
            ).value = None

            for columna in range(
                1,
                11
            ):

                celda = ws.cell(
                    fila,
                    columna
                )

                celda.fill = relleno_rojo
                celda.font = fuente_negra

            continue

        # =================================================
        # ALUMNO REGULAR
        # =================================================

        # Primero dejamos toda la fila blanca/negra.

        for columna in range(
            1,
            11
        ):

            celda = ws.cell(
                fila,
                columna
            )

            celda.fill = relleno_blanco
            celda.font = fuente_negra

        # -------------------------------------------------
        # Obtener PRM2 e INAS
        # -------------------------------------------------

        prm2 = convertir_nota(
            ws.cell(
                fila,
                columnas["PRM2"]
            ).value
        )

        inas = convertir_nota(
            ws.cell(
                fila,
                columnas["INAS"]
            ).value
        )

        # -------------------------------------------------
        # Regular excedido en faltas
        # -------------------------------------------------

        if (
            prm2 is not None
            and inas is not None
            and prm2 >= 4
            and inas > limite_faltas
        ):

            ws.cell(
                fila,
                columnas["OBSERVACION"]
            ).value = (
                "ALUMNO REGULAR EXCEDIDO EN FALTAS"
            )

            for columna in range(
                1,
                11
            ):

                ws.cell(
                    fila,
                    columna
                ).font = fuente_roja

        else:

            ws.cell(
                fila,
                columnas["OBSERVACION"]
            ).value = None

    return True

def preparar_lib(ws):
    """
    Completa la hoja LIB con los alumnos
    que tienen PRM2 menor que 4.

    Si no hay alumnos, escribe SIN ALUMNOS.

    Los alumnos libres se muestran con fondo rojo.
    """

    actualizar_progreso(
        "Preparando hoja LIB...",
        90
    )

    # --------------------------------------------------------
    # Limpiar solamente el contenido de la hoja.
    # No se elimina la hoja para conservar su formato.
    # --------------------------------------------------------

    for fila in range(
        1,
        ws.max_row + 1
    ):

        for columna in range(
            1,
            ws.max_column + 1
        ):

            ws.cell(
                fila,
                columna
            ).value = None

    # --------------------------------------------------------
    # Si no hay alumnos libres
    # --------------------------------------------------------

    if not lista_lib:

        ws.cell(
            1,
            1
        ).value = "SIN ALUMNOS"

        return True

    # --------------------------------------------------------
    # Color rojo para alumnos libres
    # --------------------------------------------------------

    relleno_rojo = PatternFill(
        fill_type="solid",
        fgColor="FF0000"
    )

    # --------------------------------------------------------
    # Encabezados
    # --------------------------------------------------------

    ws.cell(
        1,
        1
    ).value = "ALUMNO"

    ws.cell(
        1,
        2
    ).value = "PRM2"

    # --------------------------------------------------------
    # Alumnos libres
    # --------------------------------------------------------

    fila = 2

    for alumno in lista_lib:

        ws.cell(
            fila,
            1
        ).value = alumno["alumno"]

        ws.cell(
            fila,
            2
        ).value = alumno["prm2"]

        # Pintar toda la fila de rojo
        for columna in range(
            1,
            ws.max_column + 1
        ):

            ws.cell(
                fila,
                columna
            ).fill = relleno_rojo

        fila += 1

    return True

def generar_archivo_final():
    """
    Genera el archivo final completo a partir de una copia
    del archivo de inasistencias.
    """

    global archivo_salida
    global diferencias_1et

    actualizar_progreso(
        "Generando archivo final...",
        70
    )

    try:

        # ----------------------------------------------------
        # Crear copia del archivo de inasistencias
        # ----------------------------------------------------

        archivo_salida = crear_archivo_salida()

        wb = load_workbook(
            archivo_salida
        )

        # ----------------------------------------------------
        # REPORTE1 -> REPORTE
        # ----------------------------------------------------

        if "REPORTE1" in wb.sheetnames:

            if "REPORTE" in wb.sheetnames:

                del wb["REPORTE"]

            wb["REPORTE1"].title = "REPORTE"

        # ----------------------------------------------------
        # Control 1ET
        # ----------------------------------------------------
        
        validar_diferencias_1et(
            wb
        )

        # ----------------------------------------------------
        # Verificar hojas originales
        # ----------------------------------------------------

        hojas_necesarias = [
            "REPORTE",
            "INAS",
            "INAS2"
        ]

        for nombre in hojas_necesarias:

            if nombre not in wb.sheetnames:

                mostrar_error(
                    "No se encontró la hoja "
                    f"'{nombre}' en el archivo "
                    "de inasistencias."
                )

                wb.close()

                return False

        # ----------------------------------------------------
        # Eliminar LIB y REG 2ET originales
        # ----------------------------------------------------

        if "LIB" in wb.sheetnames:
            del wb["LIB"]

        if "REG 2ET" in wb.sheetnames:
            del wb["REG 2ET"]

        # ----------------------------------------------------
        # REPORTE
        # ----------------------------------------------------

        ws_reporte = wb["REPORTE"]

        # ----------------------------------------------------
        # Eliminar columna F
        # ----------------------------------------------------

        ws_reporte.delete_cols(
            6,
            1
        )

        # ----------------------------------------------------
        # Escribir notas
        # ----------------------------------------------------

        if not escribir_notas_reporte(
            ws_reporte
        ):
            wb.close()
            return False

        # ----------------------------------------------------
        # Escribir inasistencias de segunda etapa
        # ----------------------------------------------------

        if not escribir_inasistencias_reporte(
            ws_reporte
        ):
            wb.close()
            return False

        # ----------------------------------------------------
        # Marcar alumnos regulares excedidos
        # ----------------------------------------------------

        if not marcar_alumnos_excedidos_en_faltas(
            ws_reporte
        ):
            wb.close()
            return False

        # ----------------------------------------------------
        # Crear nueva hoja LIB
        # ----------------------------------------------------

        ws_lib = wb.create_sheet(
            "LIB"
        )

        if not preparar_lib(
            ws_lib
        ):
            wb.close()
            return False

        # ----------------------------------------------------
        # Trasladar color de libres a REPORTE
        # ----------------------------------------------------

        if not marcar_libres_en_reporte(
            ws_reporte
        ):
            wb.close()
            return False

        # ----------------------------------------------------
        # Guardar
        # ----------------------------------------------------

        actualizar_progreso(
            "Guardando archivo final...",
            95
        )

        wb.save(
            archivo_salida
        )

        wb.close()

        actualizar_progreso(
            "Proceso terminado correctamente.",
            100
        )

        return True

    except Exception as error:

        mostrar_error(
            "Se produjo un error al generar "
            "el archivo final:\n\n"
            f"{error}"
        )

        return False





if __name__ == "__main__":

    crear_ventana_progreso()

    if not seleccionar_archivos():
        cerrar_ventana_progreso()
        sys.exit()

    if not validar_archivos():
        cerrar_ventana_progreso()
        sys.exit()

    if not leer_datos():
        cerrar_ventana_progreso()
        sys.exit()

    if not preparar_datos_segunda_etapa():
        cerrar_ventana_progreso()
        sys.exit()

    actualizar_progreso(
        "Generando archivo final...",
        90
    )

    if not generar_archivo_final():
        cerrar_ventana_progreso()
        sys.exit()

    actualizar_progreso(
        "Proceso terminado correctamente.",
        100
    )

    # ----------------------------------------------------
    # MENSAJE DE ARCHIVO GENERADO
    # ----------------------------------------------------

    if diferencias_1et:

        messagebox.showwarning(
            TITULO,
            "Archivo generado exitosamente.\n\n"
            "ATENCIÓN: SE DETECTARON DIFERENCIAS "
            "EN 1ET.\n\n"
            "Consulte la hoja '1ET' del archivo generado."
        )

    else:

        messagebox.showinfo(
            TITULO,
            "Archivo generado exitosamente."
        )

    # ----------------------------------------------------
    # ALERTA DE APELLIDOS REPETIDOS
    # ----------------------------------------------------

    mostrar_alerta_apellidos_repetidos()