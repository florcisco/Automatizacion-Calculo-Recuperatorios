import os
import re
import copy
import queue
import threading
import tkinter as tk
import unicodedata

from tkinter import (
    filedialog,
    messagebox,
    simpledialog
)

import openpyxl

from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment
)


# ============================================================
# CONFIGURACIÓN
# ============================================================

ETAPAS = [
    "Elegir archivos",
    "Validar archivos",
    "Validar contenido",
    "Calcular",
    "Comparar",
    "Creando reporte"
]


# ============================================================
# VENTANA DE PROGRESO
# ============================================================

class VentanaProgreso:

    def __init__(self, root):

        self.root = root

        self.ventana = tk.Toplevel(root)

        self.ventana.title(
            "Procesando archivos"
        )

        self.ventana.geometry(
            "560x430"
        )

        self.ventana.resizable(
            False,
            False
        )

        self.ventana.protocol(
            "WM_DELETE_WINDOW",
            self.intentar_cerrar
        )

        titulo = tk.Label(
            self.ventana,
            text="PROCESANDO ARCHIVOS",
            font=("Arial", 16, "bold")
        )

        titulo.pack(
            pady=(25, 20)
        )

        frame_etapas = tk.Frame(
            self.ventana
        )

        frame_etapas.pack(
            fill="x",
            padx=45
        )

        self.indicadores = []
        self.textos = []

        for etapa in ETAPAS:

            frame_etapa = tk.Frame(
                frame_etapas
            )

            frame_etapa.pack(
                fill="x",
                pady=3
            )

            indicador = tk.Label(
                frame_etapa,
                text="○",
                font=("Arial", 12),
                width=3,
                anchor="w"
            )

            indicador.pack(
                side="left"
            )

            texto = tk.Label(
                frame_etapa,
                text=etapa,
                font=("Arial", 11),
                anchor="w"
            )

            texto.pack(
                side="left"
            )

            self.indicadores.append(
                indicador
            )

            self.textos.append(
                texto
            )

        self.barra_fondo = tk.Frame(
            self.ventana,
            bg="#E6E6E6",
            height=22
        )

        self.barra_fondo.pack(
            fill="x",
            padx=45,
            pady=(30, 8)
        )

        self.barra_fondo.pack_propagate(
            False
        )

        self.barra = tk.Frame(
            self.barra_fondo,
            bg="#4CAF50"
        )

        self.barra.place(
            x=0,
            y=0,
            width=0,
            relheight=1
        )

        self.porcentaje = tk.Label(
            self.ventana,
            text="0%",
            font=("Arial", 11, "bold")
        )

        self.porcentaje.pack(
            pady=(0, 10)
        )

        self.estado = tk.Label(
            self.ventana,
            text="Esperando...",
            font=("Arial", 10),
            fg="#555555"
        )

        self.estado.pack(
            pady=5
        )

        self.ventana.update_idletasks()

        ancho = self.ventana.winfo_width()
        alto = self.ventana.winfo_height()

        pantalla_ancho = (
            self.ventana.winfo_screenwidth()
        )

        pantalla_alto = (
            self.ventana.winfo_screenheight()
        )

        x = (
            pantalla_ancho - ancho
        ) // 2

        y = (
            pantalla_alto - alto
        ) // 2

        self.ventana.geometry(
            f"{ancho}x{alto}+{x}+{y}"
        )


    def actualizar(
        self,
        etapa,
        porcentaje,
        mensaje
    ):

        for i in range(
            etapa
        ):

            self.indicadores[i].config(
                text="✓",
                fg="#228B22"
            )

            self.textos[i].config(
                fg="#228B22",
                font=("Arial", 11)
            )

        if etapa < len(ETAPAS):

            self.indicadores[
                etapa
            ].config(
                text="●",
                fg="#0066CC"
            )

            self.textos[
                etapa
            ].config(
                fg="#0066CC",
                font=("Arial", 11, "bold")
            )

        for i in range(
            etapa + 1,
            len(ETAPAS)
        ):

            self.indicadores[i].config(
                text="○",
                fg="#777777"
            )

            self.textos[i].config(
                fg="#777777",
                font=("Arial", 11)
            )

        porcentaje = max(
            0,
            min(
                100,
                porcentaje
            )
        )

        self.barra.place(
            relx=0,
            rely=0,
            relwidth=porcentaje / 100,
            relheight=1
        )

        self.porcentaje.config(
            text=f"{int(porcentaje)}%"
        )

        self.estado.config(
            text=mensaje,
            fg="#555555"
        )


    def finalizar(
        self,
        mensaje
    ):

        for i in range(
            len(ETAPAS)
        ):

            self.indicadores[i].config(
                text="✓",
                fg="#228B22"
            )

            self.textos[i].config(
                fg="#228B22",
                font=("Arial", 11)
            )

        self.barra.place(
            relwidth=1
        )

        self.porcentaje.config(
            text="100%"
        )

        self.estado.config(
            text=mensaje,
            fg="#228B22",
            font=("Arial", 10, "bold")
        )


    def intentar_cerrar(self):

        messagebox.showwarning(
            "Proceso en ejecución",
            "El proceso todavía está en ejecución.\n\n"
            "Esperá a que termine antes de cerrar.",
            parent=self.ventana
        )


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def seleccionar_archivo(titulo):

    ruta = filedialog.askopenfilename(
        title=titulo,
        filetypes=[
            ("Archivos Excel", "*.xlsx *.xlsm"),
            ("Todos los archivos", "*.*")
        ]
    )

    return ruta


# ============================================================
# DNI
# ============================================================

def extraer_dni(valor):

    if valor is None:
        return None

    texto = str(valor)

    match = re.search(
        r"\bDNI\s*(\d{7,9})\b",
        texto,
        re.IGNORECASE
    )

    if match:
        return match.group(1)

    match = re.fullmatch(
        r"\s*(\d{7,9})\s*",
        texto
    )

    if match:
        return match.group(1)

    return None


# ============================================================
# NORMALIZACIÓN DE NOMBRES
# ============================================================

def normalizar_nombre(valor):

    if valor is None:
        return ""

    texto = str(
        valor
    ).strip().upper()

    # --------------------------------------------------------
    # ELIMINAR DNI
    # --------------------------------------------------------

    texto = re.sub(
        r"\bDNI\s*\d{7,9}\b",
        " ",
        texto,
        flags=re.IGNORECASE
    )

    texto = re.sub(
        r"\b\d{7,9}\b",
        " ",
        texto
    )

    # --------------------------------------------------------
    # QUITAR TILDES
    # --------------------------------------------------------

    texto = unicodedata.normalize(
        "NFD",
        texto
    )

    texto = "".join(
        caracter
        for caracter in texto
        if unicodedata.category(
            caracter
        ) != "Mn"
    )

    # --------------------------------------------------------
    # REEMPLAZAR SIGNOS POR ESPACIOS
    # --------------------------------------------------------

    texto = re.sub(
        r"[^A-Z0-9]+",
        " ",
        texto
    )

    # --------------------------------------------------------
    # NORMALIZAR ESPACIOS
    # --------------------------------------------------------

    texto = re.sub(
        r"\s+",
        " ",
        texto
    ).strip()

    return texto


# ============================================================
# NOTAS
# ============================================================

def convertir_nota(valor):

    if valor is None or valor == "":
        return 0.0

    if isinstance(
        valor,
        (int, float)
    ):

        return float(valor)

    texto = str(
        valor
    ).strip().upper()

    if texto == "A":
        return 0.0

    try:

        return float(
            texto.replace(
                ",",
                "."
            )
        )

    except ValueError:

        return 0.0


def promedio(
    nota1,
    nota2
):

    n1 = convertir_nota(
        nota1
    )

    n2 = convertir_nota(
        nota2
    )

    return (
        n1 + n2
    ) / 2


# ============================================================
# VALIDAR NOTA
# ============================================================

def validar_nota(valor):

    """
    Valida una nota individual.

    Permitidos:

        - vacío
        - A
        - entero entre 0 y 10

    No permitidos:

        - decimales
        - letras distintas de A
        - valores fuera de 0 a 10
    """

    if valor is None:
        return None

    if isinstance(
        valor,
        str
    ):

        texto = valor.strip().upper()

        if texto == "":
            return None

        if texto == "A":
            return "A"

        if not re.fullmatch(
            r"\d+",
            texto
        ):

            return (
                "formato inválido"
            )

        numero = int(
            texto
        )

    elif isinstance(
        valor,
        (int, float)
    ):

        numero_float = float(
            valor
        )

        if not numero_float.is_integer():

            return (
                "decimal"
            )

        numero = int(
            numero_float
        )

    else:

        return (
            "formato inválido"
        )

    if numero < 0 or numero > 10:

        return (
            "fuera de rango"
        )

    return numero


# ============================================================
# INASISTENCIAS
# ============================================================

def convertir_inasistencia(valor):

    if valor is None or valor == "":
        return 0.0

    try:

        return float(
            str(valor).replace(
                ",",
                "."
            )
        )

    except ValueError:

        return 0.0


def numero_limpio(valor):

    valor = float(
        valor
    )

    if valor.is_integer():
        return int(valor)

    return valor


def detectar_fila_inicio_inasistencias(ws):

    valores = []

    for columna in range(
        1,
        ws.max_column + 1
    ):

        valor = ws.cell(
            1,
            columna
        ).value

        if valor is not None:

            valores.append(
                str(valor).strip().upper()
            )

    encabezados = [
        "LEGAJO",
        "ALUMNO",
        "INASISTENCIAS",
        "INASISTENCIA",
        "FALTAS",
        "NOMBRE",
        "APELLIDO"
    ]

    for valor in valores:

        for encabezado in encabezados:

            if encabezado in valor:

                return 2

    return 1


# ============================================================
# BUSCAR COLUMNA PRM1
# ============================================================

def buscar_columna_prm1(ws):

    """
    Busca la columna cuyo encabezado sea PRM1.

    Se revisan las primeras filas de la hoja.
    """

    for fila in range(
        1,
        min(
            10,
            ws.max_row
        ) + 1
    ):

        for columna in range(
            1,
            ws.max_column + 1
        ):

            valor = ws.cell(
                fila,
                columna
            ).value

            if valor is None:
                continue

            texto = str(
                valor
            ).strip().upper()

            texto = re.sub(
                r"\s+",
                "",
                texto
            )

            if texto == "PRM1":

                return columna

    return None


# ============================================================
# PROPUESTA
# ============================================================

def contiene_propuesta(ws):

    for fila in ws.iter_rows():

        for celda in fila:

            if celda.value is None:
                continue

            valor = str(
                celda.value
            ).strip().upper()

            if valor == "PROPUESTA":

                return True

    return False


# ============================================================
# BUSCAR COLUMNA ALUMNO
# ============================================================

def buscar_columna_alumno(ws):

    palabras = [
        "ALUMNO",
        "NOMBRE",
        "APELLIDO"
    ]

    for fila in range(
        1,
        min(
            10,
            ws.max_row
        ) + 1
    ):

        for columna in range(
            1,
            ws.max_column + 1
        ):

            valor = ws.cell(
                fila,
                columna
            ).value

            if valor is None:
                continue

            texto = str(
                valor
            ).strip().upper()

            for palabra in palabras:

                if palabra in texto:

                    return columna

    if ws.max_column >= 3:

        return 3

    return None


# ============================================================
# BUSCAR FILA DE ENCABEZADOS DEL REPORTE
# ============================================================

def detectar_fila_encabezado_reporte(ws):

    """
    Busca la fila donde se encuentran los encabezados
    principales del Reporte del archivo de inasistencias.

    Se busca especialmente ALUMNO y PRM1.
    """

    columna_prm1 = buscar_columna_prm1(
        ws
    )

    if columna_prm1 is None:
        return None

    for fila in range(
        1,
        min(
            10,
            ws.max_row
        ) + 1
    ):

        tiene_alumno = False
        tiene_prm1 = False

        for columna in range(
            1,
            ws.max_column + 1
        ):

            valor = ws.cell(
                fila,
                columna
            ).value

            if valor is None:
                continue

            texto = str(
                valor
            ).strip().upper()

            texto = re.sub(
                r"\s+",
                "",
                texto
            )

            if (
                "ALUMNO" in texto
                or "NOMBRE" in texto
                or "APELLIDO" in texto
            ):

                tiene_alumno = True

            if texto == "PRM1":

                tiene_prm1 = True

        if tiene_alumno and tiene_prm1:

            return fila

    return None


# ============================================================
# OBTENER REGISTROS DE NOTAS
# ============================================================

def obtener_registros_notas(ws):

    registros = []
    duplicados = set()

    dni_vistos = {}

    for fila in range(
        11,
        ws.max_row + 1
    ):

        valor_identidad = ws.cell(
            fila,
            1
        ).value

        if (
            valor_identidad is None
            or str(valor_identidad).strip() == ""
        ):
            continue

        dni = extraer_dni(
            valor_identidad
        )

        nombre = normalizar_nombre(
            valor_identidad
        )

        if not nombre and dni is None:
            continue

        registro = {
            "dni": dni,
            "nombre": nombre,
            "fila": fila
        }

        registros.append(
            registro
        )

        if dni is not None:

            if dni in dni_vistos:

                duplicados.add(
                    dni
                )

            else:

                dni_vistos[dni] = fila

    return registros, duplicados


# ============================================================
# OBTENER REGISTROS DE INASISTENCIAS
# ============================================================

def obtener_registros_inasistencias(ws):

    """
    Obtiene alumnos únicamente de la hoja INAS o INAS2.

    IMPORTANTE:

        Esta función NO busca PRM1.

        PRM1 se obtiene exclusivamente desde la hoja
        Reporte del archivo de inasistencias.
    """

    fila_inicio = detectar_fila_inicio_inasistencias(
        ws
    )

    columna_alumno = buscar_columna_alumno(
        ws
    )

    registros = []
    duplicados = set()

    dni_vistos = {}

    for fila in range(
        fila_inicio,
        ws.max_row + 1
    ):

        legajo = ws.cell(
            fila,
            2
        ).value

        nombre = ""

        if columna_alumno is not None:

            valor_nombre = ws.cell(
                fila,
                columna_alumno
            ).value

            if valor_nombre is not None:

                nombre = normalizar_nombre(
                    valor_nombre
                )

        dni = None

        if legajo is not None:

            texto = str(
                legajo
            ).strip()

            if texto != "":

                dni = extraer_dni(
                    texto
                )

                if dni is None:

                    dni = texto

        if dni is None and not nombre:
            continue

        registro = {
            "dni": dni,
            "nombre": nombre,
            "fila": fila
        }

        registros.append(
            registro
        )

        if dni is not None:

            if dni in dni_vistos:

                duplicados.add(
                    dni
                )

            else:

                dni_vistos[dni] = fila

    return registros, duplicados


# ============================================================
# OBTENER PRM1 DESDE REPORTE DE INASISTENCIAS
# ============================================================

def obtener_registros_prm1_reporte(ws_reporte):

    """
    Obtiene los registros de PRM1 desde la hoja Reporte
    del archivo de INASISTENCIAS.

    El PRM1 se busca exclusivamente en esta hoja.

    Cada registro contiene:

        dni
        nombre
        prm1
        fila
    """

    columna_prm1 = buscar_columna_prm1(
        ws_reporte
    )

    if columna_prm1 is None:

        raise ValueError(
            "La hoja 'Reporte' del archivo de "
            "INASISTENCIAS no contiene la columna 'PRM1'."
        )

    fila_encabezado = detectar_fila_encabezado_reporte(
        ws_reporte
    )

    if fila_encabezado is None:

        fila_encabezado = 10

    columna_alumno = buscar_columna_alumno(
        ws_reporte
    )

    registros = []

    duplicados = set()

    dni_vistos = {}

    for fila in range(
        fila_encabezado + 1,
        ws_reporte.max_row + 1
    ):

        # ----------------------------------------------------
        # IDENTIDAD
        # ----------------------------------------------------

        dni = None
        nombre = ""

        # ----------------------------------------------------
        # PRIMERA OPCIÓN:
        # columna 1
        # ----------------------------------------------------

        valor_identidad = ws_reporte.cell(
            fila,
            1
        ).value

        if valor_identidad is not None:

            dni = extraer_dni(
                valor_identidad
            )

            nombre = normalizar_nombre(
                valor_identidad
            )

        # ----------------------------------------------------
        # SEGUNDA OPCIÓN:
        # columna de alumno
        # ----------------------------------------------------

        if columna_alumno is not None:

            valor_nombre = ws_reporte.cell(
                fila,
                columna_alumno
            ).value

            if valor_nombre is not None:

                nombre_desde_columna = normalizar_nombre(
                    valor_nombre
                )

                if nombre_desde_columna:

                    nombre = nombre_desde_columna

                if dni is None:

                    dni = extraer_dni(
                        valor_nombre
                    )

        # ----------------------------------------------------
        # LEGAJO / DNI EN COLUMNA 2
        # ----------------------------------------------------

        legajo = ws_reporte.cell(
            fila,
            2
        ).value

        if dni is None and legajo is not None:

            texto_legajo = str(
                legajo
            ).strip()

            if texto_legajo:

                dni = extraer_dni(
                    texto_legajo
                )

                if dni is None:

                    dni = texto_legajo

        # ----------------------------------------------------
        # PRM1
        # ----------------------------------------------------

        prm1 = ws_reporte.cell(
            fila,
            columna_prm1
        ).value

        # ----------------------------------------------------
        # IGNORAR FILAS SIN IDENTIDAD
        # ----------------------------------------------------

        if dni is None and not nombre:
            continue

        registro = {
            "dni": dni,
            "nombre": nombre,
            "prm1": prm1,
            "fila": fila
        }

        registros.append(
            registro
        )

        if dni is not None:

            if dni in dni_vistos:

                duplicados.add(
                    dni
                )

            else:

                dni_vistos[dni] = fila

    return registros, duplicados


# ============================================================
# ÍNDICE DE REGISTROS
# ============================================================

def crear_indice_dni(registros):

    indice = {}

    for registro in registros:

        dni = registro["dni"]

        if dni is None:
            continue

        indice.setdefault(
            dni,
            []
        ).append(
            registro
        )

    return indice


def crear_indice_nombres(registros):

    indice = {}

    for registro in registros:

        nombre = registro["nombre"]

        if not nombre:
            continue

        indice.setdefault(
            nombre,
            []
        ).append(
            registro
        )

    return indice


# ============================================================
# COMPARAR REGISTROS
# ============================================================

def comparar_registros(
    registros_origen,
    registros_destino
):

    faltantes = []
    coincidencias_por_nombre = []
    ambiguos = []

    indice_dni = crear_indice_dni(
        registros_destino
    )

    indice_nombres = crear_indice_nombres(
        registros_destino
    )

    for origen in registros_origen:

        dni = origen["dni"]
        nombre = origen["nombre"]

        # ----------------------------------------------------
        # CASO 1 - DNI COINCIDE
        # ----------------------------------------------------

        if (
            dni is not None
            and dni in indice_dni
        ):

            continue

        # ----------------------------------------------------
        # CASO 2 - BUSCAR POR NOMBRE
        # ----------------------------------------------------

        if nombre:

            posibles = indice_nombres.get(
                nombre,
                []
            )

            if len(posibles) == 1:

                destino = posibles[0]

                coincidencias_por_nombre.append(
                    (
                        dni,
                        destino["dni"],
                        nombre
                    )
                )

                continue

            elif len(posibles) > 1:

                dnis = sorted(
                    [
                        registro["dni"]
                        for registro in posibles
                        if registro["dni"] is not None
                    ]
                )

                ambiguos.append(
                    (
                        dni,
                        nombre,
                        dnis
                    )
                )

                continue

        # ----------------------------------------------------
        # CASO 3 - NO ENCONTRADO
        # ----------------------------------------------------

        faltantes.append(
            (
                dni,
                nombre
            )
        )

    return (
        faltantes,
        coincidencias_por_nombre,
        ambiguos
    )


# ============================================================
# FORMATEAR FALTANTES
# ============================================================

def formatear_faltantes(
    faltantes
):

    resultado = []

    for dni, nombre in faltantes:

        if nombre:

            if dni:

                resultado.append(
                    f"DNI {dni} - {nombre}"
                )

            else:

                resultado.append(
                    f"{nombre} - DNI faltante"
                )

        else:

            if dni:

                resultado.append(
                    f"DNI {dni}"
                )

            else:

                resultado.append(
                    "Alumno sin DNI y sin nombre"
                )

    return "\n".join(
        sorted(
            resultado
        )
    )


# ============================================================
# VALIDAR NOTAS DE TP3 Y TP4
# ============================================================

def validar_notas_tp3_tp4(ws_notas):

    errores = []

    for fila in range(
        11,
        ws_notas.max_row + 1
    ):

        valor_identidad = ws_notas.cell(
            fila,
            1
        ).value

        if (
            valor_identidad is None
            or str(valor_identidad).strip() == ""
        ):
            continue

        dni = extraer_dni(
            valor_identidad
        )

        nombre = normalizar_nombre(
            valor_identidad
        )

        tp3 = ws_notas.cell(
            fila,
            5
        ).value

        tp4 = ws_notas.cell(
            fila,
            6
        ).value

        resultado_tp3 = validar_nota(
            tp3
        )

        resultado_tp4 = validar_nota(
            tp4
        )

        # ----------------------------------------------------
        # AMBAS VACÍAS
        # ----------------------------------------------------

        if (
            resultado_tp3 is None
            and resultado_tp4 is None
        ):

            identificacion = nombre

            if dni:

                identificacion += (
                    f" - DNI {dni}"
                )

            errores.append(
                "TP3 y TP4 vacíos: "
                f"{identificacion}"
            )

            continue

        # ----------------------------------------------------
        # TP3
        # ----------------------------------------------------

        if isinstance(
            resultado_tp3,
            str
        ) and resultado_tp3 != "A":

            identificacion = nombre

            if dni:

                identificacion += (
                    f" - DNI {dni}"
                )

            errores.append(
                f"TP3 inválido ({tp3!r}) - "
                f"{identificacion}"
            )

        # ----------------------------------------------------
        # TP4
        # ----------------------------------------------------

        if isinstance(
            resultado_tp4,
            str
        ) and resultado_tp4 != "A":

            identificacion = nombre

            if dni:

                identificacion += (
                    f" - DNI {dni}"
                )

            errores.append(
                f"TP4 inválido ({tp4!r}) - "
                f"{identificacion}"
            )

    return errores


# ============================================================
# VALIDAR PRM1 CONTRA TP3 / TP4
# ============================================================

def validar_prm1_menor_4(
    ws_notas,
    registros_prm1
):

    """
    Valida PRM1 obtenido de la hoja Reporte del archivo
    de INASISTENCIAS contra TP3 / TP4 del archivo de NOTAS.

    Regla:

        Si PRM1 < 4:

            TP3 solamente puede ser A o vacío.
            TP4 solamente puede ser A o vacío.

    Identificación:

        1. DNI
        2. Nombre y apellido
    """

    errores = []

    # --------------------------------------------------------
    # ÍNDICES DE NOTAS
    # --------------------------------------------------------

    notas_por_dni, notas_por_nombre = crear_indice_notas(
        ws_notas
    )

    # --------------------------------------------------------
    # RECORRER PRM1
    # --------------------------------------------------------

    for registro_prm1 in registros_prm1:

        prm1 = registro_prm1.get(
            "prm1"
        )

        if prm1 is None:
            continue

        try:

            prm1_numero = float(
                str(prm1).replace(
                    ",",
                    "."
                ).strip()
            )

        except (
            ValueError,
            TypeError
        ):

            continue

        # ----------------------------------------------------
        # SOLO INTERESA PRM1 < 4
        # ----------------------------------------------------

        if prm1_numero >= 4:
            continue

        dni = registro_prm1.get(
            "dni"
        )

        nombre = registro_prm1.get(
            "nombre"
        )

        # ----------------------------------------------------
        # BUSCAR NOTAS
        # ----------------------------------------------------

        registro_notas = buscar_notas_alumno(
            dni,
            nombre,
            notas_por_dni,
            notas_por_nombre
        )

        if registro_notas is None:
            continue

        tp3 = registro_notas.get(
            "tp3"
        )

        tp4 = registro_notas.get(
            "tp4"
        )

        # ----------------------------------------------------
        # IDENTIFICACIÓN
        # ----------------------------------------------------

        identificacion = nombre

        if not identificacion:

            identificacion = (
                "Alumno sin nombre"
            )

        if dni:

            identificacion += (
                f" - DNI {dni}"
            )

        # ----------------------------------------------------
        # TP3
        # ----------------------------------------------------

        if tp3 is not None:

            texto_tp3 = str(
                tp3
            ).strip().upper()

            if texto_tp3 != "":

                if texto_tp3 != "A":

                    errores.append(
                        "PRM1 menor a 4 pero TP3 "
                        "contiene una nota distinta de A: "
                        f"TP3={tp3!r} - "
                        f"PRM1={numero_limpio(prm1_numero)} - "
                        f"{identificacion}"
                    )

        # ----------------------------------------------------
        # TP4
        # ----------------------------------------------------

        if tp4 is not None:

            texto_tp4 = str(
                tp4
            ).strip().upper()

            if texto_tp4 != "":

                if texto_tp4 != "A":

                    errores.append(
                        "PRM1 menor a 4 pero TP4 "
                        "contiene una nota distinta de A: "
                        f"TP4={tp4!r} - "
                        f"PRM1={numero_limpio(prm1_numero)} - "
                        f"{identificacion}"
                    )

    return errores


# ============================================================
# VALIDACIÓN DE CONTENIDO
# ============================================================

def validar_contenido(
    ws_notas,
    ws_reporte_inas,
    ws_inas,
    ws_inas2
):

    errores = []

    # ========================================================
    # VALIDAR TP3 Y TP4
    # ========================================================

    errores_notas = validar_notas_tp3_tp4(
        ws_notas
    )

    if errores_notas:

        errores.append(
            "ERRORES EN TP3 / TP4:\n"
            + "\n".join(
                errores_notas
            )
        )

    # ========================================================
    # OBTENER REGISTROS
    # ========================================================

    registros_notas, duplicados_notas = (
        obtener_registros_notas(
            ws_notas
        )
    )

    registros_prm1, duplicados_prm1 = (
        obtener_registros_prm1_reporte(
            ws_reporte_inas
        )
    )

    registros_inas, duplicados_inas = (
        obtener_registros_inasistencias(
            ws_inas
        )
    )

    registros_inas2, duplicados_inas2 = (
        obtener_registros_inasistencias(
            ws_inas2
        )
    )

    # ========================================================
    # VALIDAR PRM1 < 4 CONTRA TP3 / TP4
    # ========================================================

    errores_prm1 = validar_prm1_menor_4(
        ws_notas,
        registros_prm1
    )

    if errores_prm1:

        errores.append(
            "ERRORES DE PRM1 Y NOTAS:\n"
            + "\n".join(
                errores_prm1
            )
        )

    # ========================================================
    # DUPLICADOS
    # ========================================================

    if duplicados_notas:

        errores.append(
            "DNI duplicados en NOTAS:\n"
            + "\n".join(
                sorted(
                    duplicados_notas
                )
            )
        )

    if duplicados_prm1:

        errores.append(
            "DNI duplicados en REPORTE "
            "DEL ARCHIVO DE INASISTENCIAS:\n"
            + "\n".join(
                sorted(
                    duplicados_prm1
                )
            )
        )

    if duplicados_inas:

        errores.append(
            "DNI duplicados en INAS:\n"
            + "\n".join(
                sorted(
                    duplicados_inas
                )
            )
        )

    if duplicados_inas2:

        errores.append(
            "DNI duplicados en INAS2:\n"
            + "\n".join(
                sorted(
                    duplicados_inas2
                )
            )
        )

    # ========================================================
    # NOTAS VS INAS
    # ========================================================

    (
        faltantes_notas_inas,
        coincidencias_notas_inas,
        ambiguos_notas_inas
    ) = comparar_registros(
        registros_notas,
        registros_inas
    )

    # ========================================================
    # NOTAS VS INAS2
    # ========================================================

    (
        faltantes_notas_inas2,
        coincidencias_notas_inas2,
        ambiguos_notas_inas2
    ) = comparar_registros(
        registros_notas,
        registros_inas2
    )

    # ========================================================
    # INAS VS NOTAS
    # ========================================================

    (
        faltantes_inas_notas,
        coincidencias_inas_notas,
        ambiguos_inas_notas
    ) = comparar_registros(
        registros_inas,
        registros_notas
    )

    # ========================================================
    # INAS2 VS NOTAS
    # ========================================================

    (
        faltantes_inas2_notas,
        coincidencias_inas2_notas,
        ambiguos_inas2_notas
    ) = comparar_registros(
        registros_inas2,
        registros_notas
    )

    # ========================================================
    # INAS VS INAS2
    # ========================================================

    (
        faltantes_inas_inas2,
        coincidencias_inas_inas2,
        ambiguos_inas_inas2
    ) = comparar_registros(
        registros_inas,
        registros_inas2
    )

    # ========================================================
    # INAS2 VS INAS
    # ========================================================

    (
        faltantes_inas2_inas,
        coincidencias_inas2_inas,
        ambiguos_inas2_inas
    ) = comparar_registros(
        registros_inas2,
        registros_inas
    )

    # ========================================================
    # AMBIGÜEDADES
    # ========================================================

    if ambiguos_notas_inas:

        texto = []

        for dni, nombre, posibles in ambiguos_notas_inas:

            identificacion = nombre

            if dni:
                identificacion += (
                    f" - DNI {dni}"
                )

            texto.append(
                f"{identificacion}\n"
                f"Posibles DNI en INAS: "
                f"{', '.join(posibles)}"
            )

        errores.append(
            "ALUMNOS CON NOMBRE COINCIDENTE PERO "
            "MÚLTIPLES DNI POSIBLES ENTRE NOTAS E INAS:\n"
            + "\n\n".join(
                texto
            )
        )

    if ambiguos_notas_inas2:

        texto = []

        for dni, nombre, posibles in ambiguos_notas_inas2:

            identificacion = nombre

            if dni:
                identificacion += (
                    f" - DNI {dni}"
                )

            texto.append(
                f"{identificacion}\n"
                f"Posibles DNI en INAS2: "
                f"{', '.join(posibles)}"
            )

        errores.append(
            "ALUMNOS CON NOMBRE COINCIDENTE PERO "
            "MÚLTIPLES DNI POSIBLES ENTRE NOTAS E INAS2:\n"
            + "\n\n".join(
                texto
            )
        )

    if ambiguos_inas_notas:

        texto = []

        for dni, nombre, posibles in ambiguos_inas_notas:

            identificacion = nombre

            if dni:
                identificacion += (
                    f" - DNI {dni}"
                )

            texto.append(
                f"{identificacion}\n"
                f"Posibles DNI en NOTAS: "
                f"{', '.join(posibles)}"
            )

        errores.append(
            "ALUMNOS CON NOMBRE COINCIDENTE PERO "
            "MÚLTIPLES DNI POSIBLES ENTRE INAS Y NOTAS:\n"
            + "\n\n".join(
                texto
            )
        )

    if ambiguos_inas2_notas:

        texto = []

        for dni, nombre, posibles in ambiguos_inas2_notas:

            identificacion = nombre

            if dni:
                identificacion += (
                    f" - DNI {dni}"
                )

            texto.append(
                f"{identificacion}\n"
                f"Posibles DNI en NOTAS: "
                f"{', '.join(posibles)}"
            )

        errores.append(
            "ALUMNOS CON NOMBRE COINCIDENTE PERO "
            "MÚLTIPLES DNI POSIBLES ENTRE INAS2 Y NOTAS:\n"
            + "\n\n".join(
                texto
            )
        )

    # ========================================================
    # FALTANTES
    # ========================================================

    if faltantes_notas_inas:

        errores.append(
            "ALUMNOS QUE ESTÁN EN NOTAS PERO "
            "NO APARECEN EN INAS:\n"
            + formatear_faltantes(
                faltantes_notas_inas
            )
        )

    if faltantes_notas_inas2:

        errores.append(
            "ALUMNOS QUE ESTÁN EN NOTAS PERO "
            "NO APARECEN EN INAS2:\n"
            + formatear_faltantes(
                faltantes_notas_inas2
            )
        )

    if faltantes_inas_notas:

        errores.append(
            "ALUMNOS QUE ESTÁN EN INAS PERO "
            "NO APARECEN EN NOTAS:\n"
            + formatear_faltantes(
                faltantes_inas_notas
            )
        )

    if faltantes_inas2_notas:

        errores.append(
            "ALUMNOS QUE ESTÁN EN INAS2 PERO "
            "NO APARECEN EN NOTAS:\n"
            + formatear_faltantes(
                faltantes_inas2_notas
            )
        )

    if faltantes_inas_inas2:

        errores.append(
            "ALUMNOS QUE ESTÁN EN INAS PERO "
            "NO APARECEN EN INAS2:\n"
            + formatear_faltantes(
                faltantes_inas_inas2
            )
        )

    if faltantes_inas2_inas:

        errores.append(
            "ALUMNOS QUE ESTÁN EN INAS2 PERO "
            "NO APARECEN EN INAS:\n"
            + formatear_faltantes(
                faltantes_inas2_inas
            )
        )

    return errores


# ============================================================
# CREAR ÍNDICE DE NOTAS
# ============================================================

def crear_indice_notas(ws_notas):

    por_dni = {}
    por_nombre = {}

    for fila in range(
        11,
        ws_notas.max_row + 1
    ):

        valor_identidad = ws_notas.cell(
            fila,
            1
        ).value

        if (
            valor_identidad is None
            or str(valor_identidad).strip() == ""
        ):
            continue

        dni = extraer_dni(
            valor_identidad
        )

        nombre = normalizar_nombre(
            valor_identidad
        )

        tp3 = ws_notas.cell(
            fila,
            5
        ).value

        tp4 = ws_notas.cell(
            fila,
            6
        ).value

        registro = {
            "tp3": tp3,
            "tp4": tp4,
            "fila": fila,
            "dni": dni,
            "nombre": nombre
        }

        if dni is not None:

            por_dni[dni] = registro

        if nombre:

            por_nombre.setdefault(
                nombre,
                []
            ).append(
                registro
            )

    return (
        por_dni,
        por_nombre
    )


# ============================================================
# BUSCAR NOTAS DE UN ALUMNO
# ============================================================

def buscar_notas_alumno(
    dni,
    nombre,
    notas_por_dni,
    notas_por_nombre
):

    # --------------------------------------------------------
    # PRIMERO DNI
    # --------------------------------------------------------

    if (
        dni is not None
        and dni in notas_por_dni
    ):

        return notas_por_dni[dni]

    # --------------------------------------------------------
    # DESPUÉS NOMBRE
    # --------------------------------------------------------

    if nombre:

        posibles = notas_por_nombre.get(
            nombre,
            []
        )

        if len(posibles) == 1:

            return posibles[0]

    return None

# ============================================================
# COMPARAR TP1 Y TP2
# ============================================================

def comparar_tp1_tp2(
    ws_inas_reporte,
    ws_notas
):

    """
    Compara TP1 y TP2 entre:

        Archivo de INASISTENCIAS - hoja Reporte
        Archivo de NOTAS - hoja Reporte

    Columnas:

        INASISTENCIAS:
            B = TP1
            C = TP2

        NOTAS:
            B = TP1
            C = TP2

    Identificación del alumno:

        1. DNI
        2. Nombre normalizado

    Devuelve una lista de diferencias.
    """

    diferencias = []

    # --------------------------------------------------------
    # CREAR ÍNDICE DEL ARCHIVO DE NOTAS
    # --------------------------------------------------------

    registros_notas = {}

    for fila in range(
        11,
        ws_notas.max_row + 1
    ):

        valor_identidad = ws_notas.cell(
            fila,
            1
        ).value

        if (
            valor_identidad is None
            or str(valor_identidad).strip() == ""
        ):
            continue

        dni = extraer_dni(
            valor_identidad
        )

        nombre = normalizar_nombre(
            valor_identidad
        )

        registro = {
            "dni": dni,
            "nombre": nombre,
            "tp1": ws_notas.cell(
                fila,
                2
            ).value,
            "tp2": ws_notas.cell(
                fila,
                3
            ).value
        }

        # ----------------------------------------------------
        # ÍNDICE POR DNI
        # ----------------------------------------------------

        if dni is not None:

            registros_notas.setdefault(
                ("DNI", dni),
                []
            ).append(
                registro
            )

        # ----------------------------------------------------
        # ÍNDICE POR NOMBRE
        # ----------------------------------------------------

        if nombre:

            registros_notas.setdefault(
                ("NOMBRE", nombre),
                []
            ).append(
                registro
            )

    # --------------------------------------------------------
    # RECORRER INASISTENCIAS
    # --------------------------------------------------------

    for fila in range(
        11,
        ws_inas_reporte.max_row + 1
    ):

        valor_identidad = ws_inas_reporte.cell(
            fila,
            1
        ).value

        if (
            valor_identidad is None
            or str(valor_identidad).strip() == ""
        ):
            continue

        dni = extraer_dni(
            valor_identidad
        )

        nombre = normalizar_nombre(
            valor_identidad
        )

        # ----------------------------------------------------
        # BUSCAR ALUMNO EN NOTAS
        # ----------------------------------------------------

        registro_notas = None

        # Primero DNI
        if dni is not None:

            posibles = registros_notas.get(
                ("DNI", dni),
                []
            )

            if len(posibles) == 1:

                registro_notas = posibles[0]

        # Después nombre
        if registro_notas is None and nombre:

            posibles = registros_notas.get(
                ("NOMBRE", nombre),
                []
            )

            if len(posibles) == 1:

                registro_notas = posibles[0]

        # Si no se encuentra, la validación anterior
        # ya debería haber detectado el problema.
        if registro_notas is None:
            continue

        # ----------------------------------------------------
        # IDENTIFICACIÓN
        # ----------------------------------------------------

        if nombre:

            identificacion = nombre

        elif dni:

            identificacion = f"DNI {dni}"

        else:

            identificacion = "Alumno sin identificación"

        if dni:

            identificacion += f" - DNI {dni}"

        # ----------------------------------------------------
        # TP1
        # ----------------------------------------------------

        tp1_inas = ws_inas_reporte.cell(
            fila,
            2
        ).value

        tp1_notas = registro_notas[
            "tp1"
        ]

        if not notas_iguales(
            tp1_inas,
            tp1_notas
        ):

            diferencias.append(
                {
                    "alumno": identificacion,
                    "tp": "TP1",
                    "nota_controlada": tp1_inas,
                    "nota_actual": tp1_notas
                }
            )

        # ----------------------------------------------------
        # TP2
        # ----------------------------------------------------

        tp2_inas = ws_inas_reporte.cell(
            fila,
            3
        ).value

        tp2_notas = registro_notas[
            "tp2"
        ]

        if not notas_iguales(
            tp2_inas,
            tp2_notas
        ):

            diferencias.append(
                {
                    "alumno": identificacion,
                    "tp": "TP2",
                    "nota_controlada": tp2_inas,
                    "nota_actual": tp2_notas
                }
            )

    return diferencias


# ============================================================
# COMPARAR NOTAS
# ============================================================

def notas_iguales(
    nota1,
    nota2
):

    """
    Compara dos notas.

    Considera equivalentes:

        A = 0

    También permite comparar correctamente:

        vacío
        None
        0
        0.0
        "0"
        "0,0"

    Las A se consideran 0 únicamente para la comparación.
    """

    def normalizar_nota(valor):

        # ----------------------------------------------------
        # VACÍO
        # ----------------------------------------------------

        if valor is None:
            return None

        texto = str(
            valor
        ).strip().upper()

        if texto == "":
            return None

        # ----------------------------------------------------
        # A = 0
        # ----------------------------------------------------

        if texto == "A":
            return 0.0

        # ----------------------------------------------------
        # NÚMERO
        # ----------------------------------------------------

        try:

            return float(
                texto.replace(
                    ",",
                    "."
                )
            )

        except ValueError:

            return texto

    valor1 = normalizar_nota(
        nota1
    )

    valor2 = normalizar_nota(
        nota2
    )

    return valor1 == valor2


# ============================================================
# PROCESAMIENTO PRINCIPAL
# ============================================================

def procesar(
    archivo_notas,
    archivo_inas,
    limite_inasistencias,
    cola
):

    # ========================================================
    # ETAPA 2 - VALIDAR ARCHIVOS
    # ========================================================

    cola.put(
        (
            "progreso",
            1,
            15,
            "Validando archivos..."
        )
    )

    # --------------------------------------------------------
    # ABRIR ARCHIVOS
    # --------------------------------------------------------

    wb_inas = openpyxl.load_workbook(
        archivo_inas
    )

    wb_notas = openpyxl.load_workbook(
        archivo_notas,
        data_only=True
    )

    # ========================================================
    # VALIDACIÓN DE ARCHIVOS
    # ========================================================

    errores_archivos = validar_archivos(
        wb_notas,
        wb_inas
    )

    if errores_archivos:

        raise ValueError(
            "ARCHIVO INCORRECTO\n\n"
            + "\n\n".join(
                errores_archivos
            )
        )

    # --------------------------------------------------------
    # OBTENER HOJAS
    # --------------------------------------------------------

    ws_reporte = wb_inas[
        "REPORTE"
    ]

    ws_reporte_inas = wb_inas[
        "REPORTE"
    ]

    ws_inas = wb_inas[
        "INAS"
    ]

    ws_inas2 = wb_inas[
        "INAS2"
    ]

    ws_notas = wb_notas[
        "Reporte"
    ]

    # ========================================================
    # ETAPA 3 - VALIDAR CONTENIDO
    # ========================================================

    cola.put(
        (
            "progreso",
            2,
            30,
            "Validando contenido..."
        )
    )

    errores_contenido = validar_contenido(
        ws_notas,
        ws_reporte,
        ws_inas,
        ws_inas2
    )

    if errores_contenido:

        raise ValueError(
            "ERROR DE CONTENIDO\n\n"
            "Se encontraron diferencias o errores "
            "en los archivos:\n\n"
            + "\n\n".join(
                errores_contenido
            )
        )

    # ========================================================
    # ETAPA 4 - CALCULAR
    # ========================================================

    cola.put(
        (
            "progreso",
            3,
            35,
            "Cargando notas..."
        )
    )

    # ========================================================
    # ÍNDICES DE NOTAS
    # ========================================================

    notas_por_dni, notas_por_nombre = (
        crear_indice_notas(
            ws_notas
        )
    )

    # ========================================================
    # INASISTENCIAS - PRIMERA ETAPA
    # ========================================================

    inas1 = {}

    fila_inicio_inas1 = (
        detectar_fila_inicio_inasistencias(
            ws_inas
        )
    )

    for fila in range(
        fila_inicio_inas1,
        ws_inas.max_row + 1
    ):

        legajo = ws_inas.cell(
            fila,
            2
        ).value

        faltas = ws_inas.cell(
            fila,
            4
        ).value

        if legajo is None:
            continue

        dni = str(
            legajo
        ).strip()

        dni_extraido = extraer_dni(
            dni
        )

        if dni_extraido is not None:

            dni = dni_extraido

        inas1[dni] = convertir_inasistencia(
            faltas
        )

    # ========================================================
    # INASISTENCIAS - SEGUNDA ETAPA
    # ========================================================

    inas2 = {}

    fila_inicio_inas2 = (
        detectar_fila_inicio_inasistencias(
            ws_inas2
        )
    )

    for fila in range(
        fila_inicio_inas2,
        ws_inas2.max_row + 1
    ):

        legajo = ws_inas2.cell(
            fila,
            2
        ).value

        faltas = ws_inas2.cell(
            fila,
            4
        ).value

        if legajo is None:
            continue

        dni = str(
            legajo
        ).strip()

        dni_extraido = extraer_dni(
            dni
        )

        if dni_extraido is not None:

            dni = dni_extraido

        inas2[dni] = convertir_inasistencia(
            faltas
        )

    # ========================================================
    # ELIMINAR HOJAS ANTERIORES
    # ========================================================

    for nombre in [
        "LIB",
        "REG 2ET"
    ]:

        if nombre in wb_inas.sheetnames:

            del wb_inas[
                nombre
            ]

    # ========================================================
    # LIMPIAR DESDE COLUMNA E
    # ========================================================

    for fila in range(
        1,
        ws_reporte.max_row + 1
    ):

        for columna in range(
            5,
            ws_reporte.max_column + 1
        ):

            ws_reporte.cell(
                fila,
                columna
            ).value = None

    # ========================================================
    # ENCABEZADOS
    # ========================================================

    ws_reporte["E10"] = "TP3"
    ws_reporte["F10"] = "TP4"
    ws_reporte["G10"] = "PRM2"
    ws_reporte["H10"] = "INAS"
    ws_reporte["I10"] = "OBSERVACIONES"

    # ========================================================
    # COLORES
    # ========================================================

    fondo_blanco = PatternFill(
        fill_type="solid",
        fgColor="FFFFFF"
    )

    fondo_rojo = PatternFill(
        fill_type="solid",
        fgColor="FF0000"
    )

    fuente_negra = Font(
        name="Arial",
        size=10,
        color="000000"
    )

    fuente_roja = Font(
        name="Arial",
        size=10,
        color="FF0000"
    )

    # ========================================================
    # COPIAR ESTILO D -> E-I
    # ========================================================

    for columna in range(
        5,
        10
    ):

        for fila in range(
            1,
            ws_reporte.max_row + 1
        ):

            origen = ws_reporte.cell(
                fila,
                4
            )

            destino = ws_reporte.cell(
                fila,
                columna
            )

            if origen.has_style:

                destino._style = copy.copy(
                    origen._style
                )

            if origen.alignment:

                destino.alignment = copy.copy(
                    origen.alignment
                )

    # ========================================================
    # PROCESAR ALUMNOS
    # ========================================================

    alumnos_procesados = 0
    alumnos_sin_notas = 0
    alumnos_libres = 0
    alumnos_excedidos = 0

    total_filas = max(
        1,
        ws_reporte.max_row - 10
    )

    for fila in range(
        11,
        ws_reporte.max_row + 1
    ):

        porcentaje = (
            35
            + (
                (fila - 10)
                / total_filas
            ) * 25
        )

        cola.put(
            (
                "progreso",
                3,
                porcentaje,
                "Calculando resultados..."
            )
        )

        # ----------------------------------------------------
        # IDENTIDAD
        # ----------------------------------------------------

        valor_identidad = ws_reporte.cell(
            fila,
            1
        ).value

        if (
            valor_identidad is None
            or str(valor_identidad).strip() == ""
        ):
            continue

        dni = extraer_dni(
            valor_identidad
        )

        nombre = normalizar_nombre(
            valor_identidad
        )

        # ----------------------------------------------------
        # FILA BLANCA
        # ----------------------------------------------------

        for columna in range(
            1,
            10
        ):

            celda = ws_reporte.cell(
                fila,
                columna
            )

            celda.fill = copy.copy(
                fondo_blanco
            )

            celda.font = copy.copy(
                fuente_negra
            )

        # ====================================================
        # BUSCAR NOTAS
        # ====================================================

        registro_notas = buscar_notas_alumno(
            dni,
            nombre,
            notas_por_dni,
            notas_por_nombre
        )

        # ====================================================
        # TP3 / TP4 / PRM2
        # ====================================================

        if registro_notas is not None:

            tp3 = convertir_nota(
                registro_notas["tp3"]
            )

            tp4 = convertir_nota(
                registro_notas["tp4"]
            )

            prm2 = promedio(
                tp3,
                tp4
            )

            ws_reporte.cell(
                fila,
                5
            ).value = numero_limpio(
                tp3
            )

            ws_reporte.cell(
                fila,
                6
            ).value = numero_limpio(
                tp4
            )

            prm2 = numero_limpio(
                prm2
            )

            ws_reporte.cell(
                fila,
                7
            ).value = prm2

        else:

            tp3 = None
            tp4 = None
            prm2 = None

            ws_reporte.cell(
                fila,
                5
            ).value = None

            ws_reporte.cell(
                fila,
                6
            ).value = None

            ws_reporte.cell(
                fila,
                7
            ).value = None

            alumnos_sin_notas += 1

        # ====================================================
        # INASISTENCIAS
        # ====================================================

        faltas_1 = 0
        faltas_2 = 0

        if dni is not None:

            faltas_1 = inas1.get(
                dni,
                0
            )

            faltas_2 = inas2.get(
                dni,
                0
            )

        diferencia = (
            faltas_2
            - faltas_1
        )

        diferencia = numero_limpio(
            diferencia
        )

        ws_reporte.cell(
            fila,
            8
        ).value = diferencia

        # ====================================================
        # OBSERVACIONES
        # ====================================================

        ws_reporte.cell(
            fila,
            9
        ).value = None

        # ====================================================
        # ALUMNO LIBRE
        # ====================================================

        if (
            prm2 is not None
            and float(prm2) < 4
        ):

            for columna in range(
                1,
                10
            ):

                ws_reporte.cell(
                    fila,
                    columna
                ).fill = copy.copy(
                    fondo_rojo
                )

                ws_reporte.cell(
                    fila,
                    columna
                ).font = copy.copy(
                    fuente_negra
                )

            alumnos_libres += 1

        # ====================================================
        # REGULAR EXCEDIDO
        # ====================================================

        elif (
            prm2 is not None
            and float(prm2) >= 4
            and float(diferencia)
            > limite_inasistencias
        ):

            ws_reporte.cell(
                fila,
                9
            ).value = (
                "ALUMNO REGULAR EXCEDIDO EN FALTAS"
            )

            for columna in range(
                1,
                10
            ):

                ws_reporte.cell(
                    fila,
                    columna
                ).font = copy.copy(
                    fuente_roja
                )

            alumnos_excedidos += 1

        # ====================================================
        # FORMATO
        # ====================================================

        for columna in range(
            5,
            9
        ):

            ws_reporte.cell(
                fila,
                columna
            ).number_format = "General"

        alumnos_procesados += 1

    # ========================================================
    # ETAPA 5 - COMPARAR
    # ========================================================

    cola.put(
        (
            "progreso",
            4,
            65,
            "Comparando resultados..."
        )
    )

    # ========================================================
    # COMPARAR TP1 Y TP2
    # ========================================================

    diferencias_tp = comparar_tp1_tp2(
        ws_reporte_inas,
        ws_notas
    )

    # ========================================================
    # CREAR OBS
    # ========================================================

    if "OBS" in wb_inas.sheetnames:

        del wb_inas["OBS"]

    ws_obs = wb_inas.create_sheet(
        "OBS"
    )

    # --------------------------------------------------------
    # TÍTULO
    # --------------------------------------------------------

    ws_obs["A1"] = "CONTROL DE NOTAS"

    ws_obs["A1"].font = Font(
        name="Arial",
        size=12,
        bold=True
    )

    # --------------------------------------------------------
    # SIN CAMBIOS
    # --------------------------------------------------------

    if not diferencias_tp:

        ws_obs["A3"] = (
            "MATERIA SIN CAMBIOS"
        )

        ws_obs["A3"].font = Font(
            name="Arial",
            size=10,
            bold=True
        )

    else:

        # ----------------------------------------------------
        # ENCABEZADOS
        # ----------------------------------------------------

        ws_obs["A3"] = "ALUMNO"
        ws_obs["B3"] = "TP"
        ws_obs["C3"] = "NOTA CONTROLADA"
        ws_obs["D3"] = "NOTA ACTUAL"

        for columna in range(
            1,
            5
        ):

            ws_obs.cell(
                3,
                columna
            ).font = Font(
                name="Arial",
                size=10,
                bold=True
            )

        # ----------------------------------------------------
        # DIFERENCIAS
        # ----------------------------------------------------

        fila_obs = 4

        for diferencia in diferencias_tp:

            ws_obs.cell(
                fila_obs,
                1
            ).value = diferencia[
                "alumno"
            ]

            ws_obs.cell(
                fila_obs,
                2
            ).value = diferencia[
                "tp"
            ]

            ws_obs.cell(
                fila_obs,
                3
            ).value = diferencia[
                "nota_controlada"
            ]

            ws_obs.cell(
                fila_obs,
                4
            ).value = diferencia[
                "nota_actual"
            ]

            fila_obs += 1

    # --------------------------------------------------------
    # ANCHOS
    # --------------------------------------------------------

    ws_obs.column_dimensions[
        "A"
    ].width = 45

    ws_obs.column_dimensions[
        "B"
    ].width = 10

    ws_obs.column_dimensions[
        "C"
    ].width = 20

    ws_obs.column_dimensions[
        "D"
    ].width = 20
    # ========================================================
    # FORMATO GENERAL
    # ========================================================

    for fila in range(
        1,
        ws_reporte.max_row + 1
    ):

        for columna in range(
            5,
            9
        ):

            ws_reporte.cell(
                fila,
                columna
            ).number_format = "General"

    # ========================================================
    # ANCHOS
    # ========================================================

    ws_reporte.column_dimensions[
        "E"
    ].width = 12

    ws_reporte.column_dimensions[
        "F"
    ].width = 12

    ws_reporte.column_dimensions[
        "G"
    ].width = 12

    ws_reporte.column_dimensions[
        "H"
    ].width = 12

    ws_reporte.column_dimensions[
        "I"
    ].width = 42

    # ========================================================
    # ETAPA 6 - CREANDO REPORTE
    # ========================================================

    cola.put(
        (
            "progreso",
            5,
            78,
            "Creando reporte..."
        )
    )

    # ========================================================
    # CREAR LIB
    # ========================================================

    ws_lib = wb_inas.create_sheet(
        "LIB"
    )

    ws_lib["A1"] = "ALUMNOS LIBRES"

    ws_lib["A1"].font = Font(
        name="Arial",
        size=12,
        bold=True,
        color="000000"
    )

    ws_lib["A1"].alignment = Alignment(
        horizontal="center"
    )

    ws_lib["A3"] = "ALUMNO"

    ws_lib["A3"].font = Font(
        name="Arial",
        size=10,
        bold=True,
        color="000000"
    )

    # ========================================================
    # NOMBRES LIBRES
    # ========================================================

    fila_lib = 4

    for fila in range(
        11,
        ws_reporte.max_row + 1
    ):

        valor_identidad = ws_reporte.cell(
            fila,
            1
        ).value

        if (
            valor_identidad is None
            or str(valor_identidad).strip() == ""
        ):
            continue

        dni = extraer_dni(
            valor_identidad
        )

        nombre = normalizar_nombre(
            valor_identidad
        )

        registro_notas = buscar_notas_alumno(
            dni,
            nombre,
            notas_por_dni,
            notas_por_nombre
        )

        if registro_notas is None:
            continue

        prm2 = promedio(
            registro_notas["tp3"],
            registro_notas["tp4"]
        )

        if prm2 < 4:

            nombre_original = (
                ws_reporte.cell(
                    fila,
                    1
                ).value
            )

            ws_lib.cell(
                fila_lib,
                1
            ).value = nombre_original

            ws_lib.cell(
                fila_lib,
                1
            ).font = copy.copy(
                fuente_negra
            )

            fila_lib += 1

    # ========================================================
    # MENSAJE SI NO HAY ALUMNOS LIBRES
    # ========================================================

    if alumnos_libres == 0:

        ws_lib["A4"] = (
            "MATERIA SIN ALUMNOS LIBRES"
        )

        ws_lib["A4"].font = Font(
            name="Arial",
            size=10,
            bold=True,
            color="000000"
        )

        ws_lib["A4"].alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

    # ========================================================
    # FORMATO LIB
    # ========================================================

    ws_lib.column_dimensions[
        "A"
    ].width = 45

    if alumnos_libres > 0:

        for fila in range(
            4,
            fila_lib
        ):

            ws_lib.row_dimensions[
                fila
            ].height = 36

            ws_lib.cell(
                fila,
                1
            ).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True
            )

    # ========================================================
    # ARCHIVO DE SALIDA
    # ========================================================

    carpeta = os.path.dirname(
        archivo_inas
    )

    nombre = os.path.basename(
        archivo_inas
    )

    nombre_base, extension = os.path.splitext(
        nombre
    )

    archivo_salida = os.path.join(
        carpeta,
        f"{nombre_base} procesado{extension}"
    )

    # ========================================================
    # GUARDAR
    # ========================================================

    cola.put(
        (
            "progreso",
            5,
            92,
            "Guardando archivo..."
        )
    )

    wb_inas.save(
        archivo_salida
    )

    # ========================================================
    # TERMINADO
    # ========================================================

    cola.put(
        (
            "terminado",
            {
                "alumnos_procesados":
                    alumnos_procesados,

                "alumnos_libres":
                    alumnos_libres,

                "alumnos_excedidos":
                    alumnos_excedidos,

                "alumnos_sin_notas":
                    alumnos_sin_notas,

                "limite":
                    limite_inasistencias,

                "archivo_salida":
                    archivo_salida
            }
        )
    )


# ============================================================
# VALIDACIÓN DE ARCHIVOS
# ============================================================

def validar_archivos(
    wb_notas,
    wb_inas
):

    errores = []

    # --------------------------------------------------------
    # ARCHIVO DE NOTAS
    # --------------------------------------------------------

    if "Reporte" not in wb_notas.sheetnames:

        errores.append(
            "El archivo de NOTAS no contiene "
            "la hoja 'Reporte'."
        )

    else:

        ws_notas = wb_notas[
            "Reporte"
        ]

        if not contiene_propuesta(
            ws_notas
        ):

            errores.append(
                "El archivo de NOTAS no contiene "
                "la columna o encabezado 'PROPUESTA'."
            )

    # --------------------------------------------------------
    # ARCHIVO DE INASISTENCIAS
    # --------------------------------------------------------

    if "INAS" not in wb_inas.sheetnames:

        errores.append(
            "El archivo de INASISTENCIAS no contiene "
            "la hoja 'INAS'."
        )

    if "INAS2" not in wb_inas.sheetnames:

        errores.append(
            "El archivo de INASISTENCIAS no contiene "
            "la hoja 'INAS2'."
        )

    # --------------------------------------------------------
    # REPORTE DEL ARCHIVO DE INASISTENCIAS
    # --------------------------------------------------------

    if "REPORTE" not in wb_inas.sheetnames:

        errores.append(
            "El archivo de INASISTENCIAS no contiene "
            "la hoja 'Reporte'."
        )

    else:

        ws_reporte = wb_inas[
            "REPORTE"
        ]

        if buscar_columna_prm1(
            ws_reporte
        ) is None:

            errores.append(
                "La hoja 'Reporte' del archivo de "
                "INASISTENCIAS no contiene "
                "la columna 'PRM1'."
            )

    return errores


# ============================================================
# INTERFAZ PRINCIPAL
# ============================================================

def main():

    root = tk.Tk()

    root.withdraw()

    # ========================================================
    # COLA DE COMUNICACIÓN
    # ========================================================

    cola = queue.Queue()

    # ========================================================
    # VENTANA DE PROGRESO
    # ========================================================

    progreso = VentanaProgreso(
        root
    )

    # ========================================================
    # SELECCIÓN DE ARCHIVOS
    # ========================================================

    progreso.actualizar(
        0,
        0,
        "Seleccioná el archivo de NOTAS..."
    )

    messagebox.showinfo(
        "Paso 1 de 3",
        "Primero seleccioná el archivo de NOTAS.",
        parent=progreso.ventana
    )

    archivo_notas = seleccionar_archivo(
        "Seleccionar archivo de NOTAS"
    )

    if not archivo_notas:

        progreso.ventana.destroy()
        root.destroy()

        return

    progreso.actualizar(
        0,
        5,
        "Seleccioná el archivo de INASISTENCIAS..."
    )

    messagebox.showinfo(
        "Paso 2 de 3",
        "Ahora seleccioná el archivo de INASISTENCIAS.",
        parent=progreso.ventana
    )

    archivo_inas = seleccionar_archivo(
        "Seleccionar archivo de INASISTENCIAS"
    )

    if not archivo_inas:

        progreso.ventana.destroy()
        root.destroy()

        return

    progreso.actualizar(
        0,
        10,
        "Indicá el límite de inasistencias..."
    )

    limite = simpledialog.askfloat(
        "Paso 3 de 3",
        "¿Cuál es el límite de inasistencias?",
        parent=progreso.ventana,
        minvalue=0
    )

    if limite is None:

        progreso.ventana.destroy()
        root.destroy()

        return

    # ========================================================
    # ELEGIR ARCHIVOS TERMINADO
    # ========================================================

    progreso.actualizar(
        0,
        15,
        "Archivos seleccionados. Iniciando procesamiento..."
    )

    # ========================================================
    # PROCESAMIENTO EN SEGUNDO PLANO
    # ========================================================

    def ejecutar():

        try:

            procesar(
                archivo_notas,
                archivo_inas,
                limite,
                cola
            )

        except Exception as e:

            cola.put(
                (
                    "error",
                    type(e).__name__,
                    str(e)
                )
            )

    hilo = threading.Thread(
        target=ejecutar,
        daemon=True
    )

    hilo.start()

    # ========================================================
    # REVISAR COLA
    # ========================================================

    def revisar_cola():

        try:

            while True:

                mensaje = cola.get_nowait()

                # --------------------------------------------
                # ACTUALIZAR PROGRESO
                # --------------------------------------------

                if mensaje[0] == "progreso":

                    _, etapa, porcentaje, texto = mensaje

                    progreso.actualizar(
                        etapa,
                        porcentaje,
                        texto
                    )

                # --------------------------------------------
                # TERMINADO
                # --------------------------------------------

                elif mensaje[0] == "terminado":

                    resultado = mensaje[1]

                    progreso.finalizar(
                        "Proceso terminado correctamente."
                    )

                    messagebox.showinfo(
                        "Proceso terminado",

                        "El archivo fue procesado correctamente.\n\n"

                        f"Alumnos procesados: "
                        f"{resultado['alumnos_procesados']}\n"

                        f"Alumnos libres: "
                        f"{resultado['alumnos_libres']}\n"

                        f"Regulares excedidos en faltas: "
                        f"{resultado['alumnos_excedidos']}\n"

                        f"Alumnos sin notas: "
                        f"{resultado['alumnos_sin_notas']}\n\n"

                        f"Límite de inasistencias: "
                        f"{resultado['limite']}\n\n"

                        "Archivo generado:\n"
                        f"{resultado['archivo_salida']}",

                        parent=progreso.ventana
                    )

                    progreso.ventana.destroy()
                    root.destroy()

                    return

                # --------------------------------------------
                # ERROR
                # --------------------------------------------

                elif mensaje[0] == "error":

                    _, tipo, detalle = mensaje

                    messagebox.showerror(
                        "Error",

                        "Ocurrió un error durante "
                        "el procesamiento:\n\n"

                        f"{tipo}: {detalle}",

                        parent=progreso.ventana
                    )

                    progreso.ventana.destroy()
                    root.destroy()

                    return

        except queue.Empty:

            pass

        root.after(
            100,
            revisar_cola
        )

    # ========================================================
    # INICIAR CONTROL DE COLA
    # ========================================================

    root.after(
        100,
        revisar_cola
    )

    # ========================================================
    # LOOP PRINCIPAL
    # ========================================================

    root.mainloop()


# ============================================================
# EJECUTAR
# ============================================================

if __name__ == "__main__":
    main()
